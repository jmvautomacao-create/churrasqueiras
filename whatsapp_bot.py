import asyncio
import json
import os
import time
import re
import unicodedata
import httpx
from pathlib import Path
from datetime import datetime
from playwright.async_api import async_playwright

from config import PRODUTOS, SEU_NUMERO, TRANSPORTADORAS, BASE_DIR, PASTA_SOLICITACOES
from database import (
    cliente_por_telefone, criar_cliente, criar_conversa, salvar_mensagem,
    atualizar_etapa_conversa, atualizar_produto_interesse, criar_cotacao,
    atualizar_cotacao, criar_venda, get_historico_conversa, get_conversa_ativa,
    atualizar_cliente, get_ultima_cotacao, confirmar_pagamento, get_venda,
    get_venda_pendente_conversa, tem_compra_finalizada,
)
from gemini_agent import gerar_resposta, resposta_fallback, extrair_comando, limpar_resposta
from stripe_integration import criar_checkout_pix_cartao, verificar_pagamento
from produtos import valor_produto, produto_por_id, detalhar


def safe(texto):
    if not isinstance(texto, str):
        texto = str(texto)
    return texto.encode('ascii', errors='replace').decode('ascii')


class WhatsAppBot:
    @staticmethod
    def _n(nome: str) -> str:
        nome = unicodedata.normalize('NFKC', nome)
        nome = re.sub(r'\s+', ' ', nome).strip()
        return nome

    @staticmethod
    def _strip_emoji(texto: str) -> str:
        return re.sub(r'[^\x00-\x7F\s]', '', texto)

    def __init__(self):
        self.page = None
        self.context = None
        self.playwright = None
        self.logado = False
        self.processando: dict[str, bool] = {}
        self.ultimo_processamento: dict[str, float] = {}
        self.ultimo_texto_chat: dict[str, str] = {}
        self.ultimo_visto_texto: dict[str, float] = {}
        self.ultimo_envio: dict[str, float] = {}
        self.ultimo_envio_texto: dict[str, str] = {}
        self.mapa_contatos = {}
        self.ultimo_mapa = 0
        self.apresentacao_menu: dict[str, dict] = {}
        self.apresentacao_submenu: dict[str, dict] = {}
        self.continuar_submenu: dict[str, dict] = {}
        self.primeiro_ciclo = True
        self.ultimo_fallback: dict[str, float] = {}
        self.ultima_limpeza = 0.0
        self.chats_com_resposta: set[str] = set()
        self.ultimo_gemini: dict[str, float] = {}
        self.fretes_pendentes: dict[str, dict] = {}
        self.proximo_id_frete: int = 0
        self._respostas_frete_vistas: set[str] = set()
        self._cache_cep: dict[str, dict] = {}
        self._cache_endereco: dict[str, dict] = {}
        self.fila_mensagens: asyncio.Queue | None = None
        self.fila_pendentes: set[str] = set()
        self.fila_worker_task: asyncio.Task | None = None
        self.frete_monitor_task: asyncio.Task | None = None
        self.sidebar_lock = asyncio.Lock()
        self._pagina_perdida = False
        self.ultimo_envio_sucesso: dict[str, float] = {}
        self.transport_page = None
        self.transport_ultimo_texto_sidebar: dict[str, str] = {}
        self.transport_ultimo_id_msg: dict[str, str] = {}

    async def iniciar(self):
        self.playwright = await async_playwright().start()
        user_data_dir = str(BASE_DIR / "data" / "whatsapp_session")

        sessao_dir = Path(user_data_dir)
        sessao_dir.mkdir(parents=True, exist_ok=True)
        print(f"Sessão: {user_data_dir}")

        self.context = await self.playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir, headless=False,
            args=['--window-position=0,0', '--window-size=1280,900'],
        )
        # Reusar página existente (evita janela extra duplicada de sessão anterior)
        pages = self.context.pages
        self.page = pages[0] if pages else await self.context.new_page()
        await self.page.goto("https://web.whatsapp.com")
        # Trazer janela para o foco e garantir posição no monitor principal
        try:
            await self.avaliar("() => { window.focus(); window.moveTo(0, 0); window.resizeTo(1280, 900); }")
        except:
            pass
        print("Aguardando login. Se já estiver logado, isso leva segundos.")

        for i in range(120):
            await asyncio.sleep(2)
            try:
                url_atual = self.page.url
                titulo = await self.page.title()
                print(f"  [{i*2}s] {safe(titulo[:40])} | {safe(url_atual[:50])}")

                if "web.whatsapp.com" not in url_atual:
                    print("  -> Redirecionado, reabrindo web.whatsapp.com...")
                    await self.page.goto("https://web.whatsapp.com")
                    await asyncio.sleep(2)
                    continue

                panel = await self.page.query_selector('[data-testid="conversation-panel-main"]')
                if panel:
                    print("Login confirmado!")
                    self.logado = True
                    self.fila_mensagens = asyncio.Queue()
                    self.fila_worker_task = asyncio.create_task(self._worker())
                    self.frete_monitor_task = asyncio.create_task(self._monitorar_fretes())
                    await self._iniciar_transport_page()
                    await asyncio.sleep(2)
                    return True

                side = await self.page.query_selector('#side')
                if side:
                    print("Sidebar detectada - logado!")
                    self.logado = True
                    self.fila_mensagens = asyncio.Queue()
                    self.fila_worker_task = asyncio.create_task(self._worker())
                    self.frete_monitor_task = asyncio.create_task(self._monitorar_fretes())
                    await self._iniciar_transport_page()
                    await asyncio.sleep(2)
                    return True
            except Exception as e:
                print(f"  -> Erro: {safe(e)}")
                await asyncio.sleep(2)

        print("Tempo limite excedido.")
        return False

    async def _iniciar_transport_page(self):
        try:
            self.transport_page = await self.context.new_page()
            await self.transport_page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(3)
            print("  [transport] Página criada para monitoramento de fretes", flush=True)
        except Exception as e:
            print(f"  [transport] Erro ao criar página: {safe(str(e)[:80])}", flush=True)
            self.transport_page = None

    async def avaliar(self, codigo: str, *args):
        try:
            return await self.page.evaluate(codigo, *args)
        except Exception as e:
            msg = str(e).lower()
            if "context" in msg or "navigation" in msg or "target closed" in msg:
                self._pagina_perdida = True
                print(f"  -> Página perdida durante evaluate: {safe(str(e)[:60])}")
            raise

    async def _recuperar_pagina(self):
        print("  -> Recuperando página...")
        await asyncio.sleep(2)
        try:
            if not self.page or self.page.is_closed():
                self.page = await self.context.new_page()
            await self.page.goto("https://web.whatsapp.com", wait_until="load", timeout=30000)
        except Exception as e:
            print(f"  -> Erro na recuperação: {safe(str(e)[:60])}, tentando nova página...")
            try:
                self.page = await self.context.new_page()
                await self.page.goto("https://web.whatsapp.com", wait_until="load", timeout=30000)
            except:
                pass
        for i in range(15):
            try:
                if await self.page.query_selector('#side'):
                    print("  -> Login confirmado após recuperação.")
                    await asyncio.sleep(2)
                    return
            except:
                pass
            await asyncio.sleep(1)
        print("  -> Página carregada (pode precisar de QR code).")
        await asyncio.sleep(2)

    async def _garantir_pagina_principal(self):
        try:
            url = self.page.url
            if "web.whatsapp.com" not in url:
                print("  -> Fora do WhatsApp, navegando...")
                await self.page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=15000)
                await asyncio.sleep(3)
        except:
            pass

    async def _perguntar_nova_compra_ou_info(
        self, telefone: str, conv_id: int, nome_cliente: str,
        *, alterar_etapa: bool = False
    ) -> None:
        if alterar_etapa:
            atualizar_etapa_conversa(conv_id, "pos_compra")
        msg = (
            f'Olá! Seja Bem-vindo "{nome_cliente}", '
            f"digite um número das opções abaixo:\n\n"
            f"[1] Nova compra\n"
            f"[2] Informações da compra anterior"
        )
        await self.enviar_para_cliente(telefone, msg)
        salvar_mensagem(conv_id, "agente", "Nova compra ou informações?")
        self.processando.pop(telefone, None)

    async def _iniciar_apresentacao_menu(self, telefone: str, conv_id: int, nome_sidebar: str = "") -> bool:
        chaves = [k for k in self.ultimo_visto_texto if k.startswith(f"{telefone}|")]
        for k in chaves:
            self.ultimo_visto_texto.pop(k, None)
        from produtos import menu_interativo
        texto = menu_interativo()
        ok = await self.enviar_para_cliente(telefone, texto, nome_sidebar)
        if not ok:
            return False
        salvar_mensagem(conv_id, "agente", texto)
        self.apresentacao_menu[telefone] = {
            "conv_id": conv_id,
            "apresentados": [p["id"] for p in PRODUTOS],
            "todos_enviados": True,
        }
        atualizar_etapa_conversa(conv_id, "apresentacao_menu")
        return True

    async def _avancar_apresentacao_menu(self, telefone: str, conv_id: int, estado: dict):
        pass

    async def _iniciar_apresentacao_submenu(self, telefone: str, conv_id: int, produto: dict):
        chaves = [k for k in self.ultimo_visto_texto if k.startswith(f"{telefone}|")]
        for k in chaves:
            self.ultimo_visto_texto.pop(k, None)
        from produtos import submenu_produto
        texto = submenu_produto(produto)
        await self.enviar_para_cliente(telefone, texto)
        salvar_mensagem(conv_id, "agente", texto)
        self.apresentacao_submenu[telefone] = {
            "conv_id": conv_id,
            "produto_id": produto["id"],
            "apresentados": [1, 3, 4, 5, 6],
            "todos_enviados": True,
        }
        atualizar_etapa_conversa(conv_id, "apresentacao_submenu")

    async def _avancar_apresentacao_submenu(self, telefone: str, conv_id: int, estado: dict):
        pass

    async def _executar_opcao_submenu(self, telefone: str, conv_id: int, produto_id: int, opt: int, nome_sidebar: str = ""):
        produto = produto_por_id(produto_id)
        if not produto:
            return
        if opt == 1:
            await self._enviar_folder(conv_id, telefone, produto)
        elif opt == 2:
            resp = valor_produto(produto)
            await self.enviar_para_cliente(telefone, resp)
            salvar_mensagem(conv_id, "agente", resp)
        elif opt == 3:
            await self._enviar_foto(conv_id, telefone, produto)
        elif opt == 4:
            await self._enviar_video(conv_id, telefone, produto)
        elif opt == 5:
            atualizar_produto_interesse(conv_id, produto["id"])
            from database import cliente_por_telefone
            cli = cliente_por_telefone(telefone)
            if cli and all([cli.get("nome"), cli.get("cpf_cnpj"), cli.get("endereco"), cli.get("cidade"), cli.get("estado"), cli.get("cep")]):
                await self._finalizar_coleta_frete(conv_id, cli["id"], telefone, nome_sidebar)
                return
            atualizar_etapa_conversa(conv_id, "frete_nome")
            await self.enviar_para_cliente(telefone,
                f"Para solicitar o frete da {produto['nome']}, preciso de alguns dados.\n\n"
                f"Primeiro, informe seu NOME completo:")
            salvar_mensagem(conv_id, "agente", "Solicitando dados para frete - informe o nome:")
            return
        elif opt == 6:
            self.apresentacao_submenu.pop(telefone, None)
            await self.enviar_para_cliente(telefone, "Voltando ao Menu Principal...", nome_sidebar)
            ok = await self._iniciar_apresentacao_menu(telefone, conv_id, nome_sidebar)
            if ok:
                print(f"  -> Menu reiniciado para {safe(telefone)}", flush=True)
            return
        chaves = [k for k in self.ultimo_visto_texto if k.startswith(f"{telefone}|")]
        for k in chaves:
            self.ultimo_visto_texto.pop(k, None)
        self.continuar_submenu[telefone] = {"conv_id": conv_id, "produto_id": produto_id, "nome_sidebar": nome_sidebar}
        atualizar_etapa_conversa(conv_id, "submenu_continuar")
        await self.enviar_para_cliente(telefone,
            "Selecione uma opção abaixo:\n[f] Continuar neste produto\n[g] Voltar ao Menu Principal",
            nome_sidebar)

    async def _atualizar_mapa_contatos(self):
        agora = time.time()
        if agora - self.ultimo_mapa < 30:
            return
        try:
            raw = await self.avaliar("""
                async (selfNum) => {
                    const db = await new Promise(r => {
                        const req = indexedDB.open('model-storage');
                        req.onsuccess = () => r(req.result);
                    });
                    const out = {};

                    // 1. Contact store: extrai telefone apenas de JIDs com dominio c.us ou s.whatsapp.net
                    {
                        const tx = db.transaction('contact', 'readonly');
                        const store = tx.objectStore('contact');
                        const all = await new Promise(r => {
                            const req = store.getAll();
                            req.onsuccess = () => r(req.result);
                        });
                        for (const c of all) {
                            const name = c.name || c.pushname || '';
                            if (!name || !c.id) continue;
                            const parts = c.id.split('@');
                            const domain = parts[1] || '';
                            const phone = parts[0];
                            if ((domain === 'c.us' || domain === 's.whatsapp.net') && /^\\d+$/.test(phone) && phone !== selfNum) {
                                out[name] = phone;
                            }
                        }
                    }

                    // 2. Message store: lid -> phone, cruzar com contact names
                    {
                        const lidPhone = {};
                        const tx = db.transaction('message', 'readonly');
                        const store = tx.objectStore('message');
                        const all = await new Promise(r => {
                            const req = store.getAll();
                            req.onsuccess = () => r(req.result);
                        });
                        for (const m of all) {
                            if (!m.id) continue;
                            const parts = m.id.split('_');
                            if (parts.length < 2) continue;
                            const lid = parts[1];
                            if (lidPhone[lid]) continue;
                            let phone = '';
                            if (m.from && typeof m.from === 'string') {
                                const match = m.from.match(/^(\\d+)@/);
                                if (match && match[1] !== selfNum) phone = match[1];
                            }
                            if (!phone && m.to) {
                                const toUser = typeof m.to === 'object' ? (m.to.user || '') :
                                               (typeof m.to === 'string' ? m.to.split('@')[0] : '');
                                if (/^\\d+$/.test(toUser) && toUser !== selfNum) phone = toUser;
                            }
                            if (phone) lidPhone[lid] = phone;
                        }
                        const tx2 = db.transaction('contact', 'readonly');
                        const store2 = tx2.objectStore('contact');
                        const all2 = await new Promise(r => {
                            const req = store2.getAll();
                            req.onsuccess = () => r(req.result);
                        });
                        for (const c of all2) {
                            const name = c.name || c.pushname || '';
                            if (!name || !c.id) continue;
                            const phone = lidPhone[c.id];
                            if (phone && !out[name]) out[name] = phone;
                        }
                    }

                    return JSON.stringify(out);
                }
            """, SEU_NUMERO)
            self.mapa_contatos = json.loads(raw)
            self.ultimo_mapa = agora
            print(f"  [mapa] {len(self.mapa_contatos)} contatos mapeados")
        except Exception as e:
            print(f"  [mapa] erro: {safe(str(e)[:80])}")

    async def _limpar_dicts_antigos(self):
        agora = time.time()
        if agora - self.ultima_limpeza < 3600:
            return
        self.ultima_limpeza = agora
        limite = 7200
        for k in list(self.ultimo_visto_texto):
            if agora - self.ultimo_visto_texto[k] > limite:
                del self.ultimo_visto_texto[k]
        for k in list(self.ultimo_envio):
            if agora - self.ultimo_envio[k] > limite:
                del self.ultimo_envio[k]
        for k in list(self.ultimo_envio_texto):
            if agora - self.ultimo_envio.get(k, 0) > limite:
                del self.ultimo_envio_texto[k]
        for k in list(self.ultimo_fallback):
            if agora - self.ultimo_fallback[k] > limite:
                del self.ultimo_fallback[k]
        for k in list(self.ultimo_gemini):
            if agora - self.ultimo_gemini[k] > limite:
                del self.ultimo_gemini[k]
        for k in list(self.ultimo_processamento):
            if agora - self.ultimo_processamento[k] > limite:
                del self.ultimo_processamento[k]
        for k in list(self.processando):
            if agora - self.ultimo_processamento.get(k, 0) > 300:
                del self.processando[k]
        # Limpa travamentos na fila de atendimento
        for k in list(self.fila_pendentes):
            if agora - self.ultimo_processamento.get(k, 0) > 600:
                self.fila_pendentes.discard(k)
        for k in list(self.ultimo_envio_sucesso):
            if agora - self.ultimo_envio_sucesso[k] > limite:
                del self.ultimo_envio_sucesso[k]
        # Limpa dedup de respostas de frete (retem apenas ultima 1h)
        self._respostas_frete_vistas.clear()
        # Limpa cache de IDs de mensagens de transportadora
        for k in list(self.transport_ultimo_id_msg):
            if agora - self.ultimo_processamento.get(k, 0) > limite:
                del self.transport_ultimo_id_msg[k]
        for k in list(self.transport_ultimo_texto_sidebar):
            if agora - self.ultimo_processamento.get(k, 0) > limite:
                del self.transport_ultimo_texto_sidebar[k]
        print(f"  -> Dicts limpos (retidos {limite//3600}h)", flush=True)

    async def detectar_chats(self):
        await self._garantir_pagina_principal()
        await self._atualizar_mapa_contatos()
        mapa_str = json.dumps(self.mapa_contatos)
        codigo = """
            (mapaStr) => {
                const mapa = JSON.parse(mapaStr);
                const achados = [];
                const side = document.querySelector('#side') ||
                             document.querySelector('[role="tabpanel"]');
                if (!side) {
                    const rows = document.querySelectorAll('[role="row"]');
                    if (rows.length > 0) {
                        rows.forEach((row, i) => {
                            if (i > 30) return;
                            const el = row.querySelector('[title]');
                            const nome = el ? el.getAttribute('title') : '';
                            if (nome && nome.length < 30 && nome !== 'Filtrar conversas' && !nome.startsWith('Filt')) {
                                const spans = row.querySelectorAll('span[dir]');
                                let texto = '';
                                const titulo = el ? el.getAttribute('title') : '';
                                for (const sp of spans) {
                                    if (sp.getAttribute('title') !== titulo && sp.textContent.trim()) {
                                        texto = sp.textContent.trim();
                                    }
                                }
                                if (texto.startsWith('default-')) return;
                                const badge = row.querySelector('[data-testid="icon-unread-count"]') ||
                                             row.querySelector('[aria-label*="nao lida"]') ||
                                             row.querySelector('[aria-label*="unread"]');
                                let telefone = nome.replace(/\\D/g, '');
                                if (!telefone || telefone.length < 10) {
                                    telefone = mapa[nome] || '';
                                }
                                if (!texto && !telefone) return;
                                achados.push({nome, texto, nao_lida: !!badge, telefone});
                            }
                        });
                        return JSON.stringify(achados);
                    }
                    return '[]';
                }

                const chats = side.querySelectorAll(':scope [role="row"]');
                chats.forEach(chat => {
                    const el = chat.querySelector('[title]');
                    const nome = el ? el.getAttribute('title') : '';
                    if (!nome || nome.length > 30 || nome === 'Filtrar conversas' || nome.startsWith('Filt')) return;

                    const spans = chat.querySelectorAll('span[dir]');
                    let texto = '';
                    const titulo = el ? el.getAttribute('title') : '';
                    for (const sp of spans) {
                        if (sp.getAttribute('title') !== titulo && sp.textContent.trim()) {
                            texto = sp.textContent.trim();
                        }
                    }
                    if (texto.startsWith('default-')) return;

                    const badge = chat.querySelector('[data-testid="icon-unread-count"]') ||
                                 chat.querySelector('[aria-label*="nao lida"]') ||
                                 chat.querySelector('[aria-label*="unread"]');

                    let telefone = nome.replace(/\\D/g, '');
                    if (!telefone || telefone.length < 10) {
                        telefone = mapa[nome] || '';
                    }

                    if (!texto && !telefone) return;

                    achados.push({nome, texto, nao_lida: !!badge, telefone});
                });
                return JSON.stringify(achados);
            }
        """
        return await self.avaliar(codigo, mapa_str)

    async def _monitorar_fretes(self):
        while True:
            try:
                if self.fretes_pendentes:
                    await self._transport_processar_fretes()
                await self._verificar_pagamentos_pendentes()
                await asyncio.sleep(5)
            except asyncio.CancelledError:
                break
            except Exception as e:
                print(f"[FRETE MONITOR] {safe(e)}", flush=True)
                import traceback
                traceback.print_exc()
                await asyncio.sleep(10)

    async def _verificar_pagamentos_pendentes(self):
        try:
            from database import get_connection
            from stripe_integration import verificar_pagamento
            conn = get_connection()
            rows = conn.execute(
                "SELECT id, conversa_id, cliente_id, stripe_session_id, xlsx_path FROM vendas WHERE payment_status='pendente' AND stripe_session_id IS NOT NULL"
            ).fetchall()
            conn.close()
            for row in rows:
                venda = dict(row)
                sess_id = venda["stripe_session_id"]
                if not verificar_pagamento(sess_id):
                    continue
                conn2 = get_connection()
                conn2.execute("UPDATE vendas SET payment_status='pago', status='pago' WHERE id=?", (venda["id"],))
                conn2.execute("UPDATE conversas SET etapa='fechada' WHERE id=?", (venda["conversa_id"],))
                conn2.commit()
                cli = conn2.execute("SELECT telefone, nome FROM clientes WHERE id=?", (venda["cliente_id"],)).fetchone()
                conn2.close()
                if cli:
                    msg_confirm = f"✅ Pagamento confirmado! Seu pedido será processado em breve. Obrigado, {cli['nome']}!"
                    await self.enviar_para_cliente(cli["telefone"], msg_confirm)
                    # Evita que essa mensagem seja detectada como nova mensagem do usuario
                    self.ultimo_visto_texto[f"{cli['telefone']}|{self._n(msg_confirm)}"] = time.time() + 3600
                    await self.enviar_para_cliente(SEU_NUMERO,
                        f"✅ PAGAMENTO CONFIRMADO (auto) - Venda #{venda['id']} - {cli['nome']}")
                    self._atualizar_xlsx_venda_paga(venda.get("xlsx_path", ""))
                    print(f"[PAGAMENTO AUTO] Venda {venda['id']} confirmada via Stripe", flush=True)
        except Exception as e:
            print(f"[VERIFICAR PAGAMENTOS] {safe(e)}", flush=True)

    async def _worker(self):
        while True:
            try:
                item = await self.fila_mensagens.get()
                nome_key, texto, telefone, nome_raw = item
                try:
                    print(f"\n>>> [WORKER] Processando {safe(nome_raw)}: {safe(texto)}", flush=True)
                    if texto:
                        self.ultimo_visto_texto[f"{telefone}|{self._n(texto)}"] = time.time()
                    await self.processar_mensagem(nome_key, texto, telefone, nome_raw)
                except Exception as e:
                    print(f"[WORKER ERRO] {safe(e)}", flush=True)
                    import traceback
                    traceback.print_exc()
                finally:
                    if texto:
                        self.ultimo_visto_texto[f"{telefone}|{self._n(texto)}"] = time.time()
                    self.fila_pendentes.discard(telefone)
                    self.fila_mensagens.task_done()
            except asyncio.CancelledError:
                break
            except Exception as e:
                print(f"[WORKER FATAL] {safe(e)}", flush=True)
                import traceback
                traceback.print_exc()
                await asyncio.sleep(5)

    async def escutar_mensagens(self):
        print("\n" + "="*50)
        print("Ouvindo mensagens... Ctrl+C para parar.")
        print("="*50 + "\n")

        c = 0
        erros_consecutivos = 0
        while True:
            try:
                c += 1
                raw = await self.detectar_chats()
                chats = json.loads(raw)
                erros_consecutivos = 0

                if c % 10 == 0:
                    queue_size = self.fila_mensagens.qsize() if self.fila_mensagens else 0
                    pendentes = len(self.fila_pendentes)
                    print(f"  [{c}] heartbeat - {len(chats)} chats, fila: {queue_size}, pendentes: {pendentes}, menu: {len(self.apresentacao_menu)}, submenu: {len(self.apresentacao_submenu)}")

                vistos_ciclo = set()
                for chat in chats:
                    nome_raw = chat.get("nome", "")
                    nome_key = self._n(nome_raw) if nome_raw else ""
                    texto = chat.get("texto", "")
                    nao_lida = chat.get("nao_lida", False)
                    telefone = chat.get("telefone", "")

                    if not nome_raw or nome_raw == "DEBUG" or nome_raw.startswith("Filt"):
                        continue

                    # Transportadoras: ignorar no loop principal (processadas via transport_page/IndexedDB)
                    transportadoras_tels = set(t["numero"] for t in TRANSPORTADORAS) | {self.TRANSPORTADORA_FOB}
                    if telefone in transportadoras_tels:
                        if c % 30 == 0:
                            print(f"  [{c} SKIP] {safe(nome_raw)}: transportadora ignorada no loop principal")
                        continue

                    # Dedup intra-ciclo: chats duplicados (ex: role="row" aninhado)
                    if nome_key in vistos_ciclo:
                        continue
                    vistos_ciclo.add(nome_key)

                    if not telefone or len(telefone) < 12:
                        continue

                    if c % 30 == 0 or nao_lida:
                        marca = f" [tel:{telefone}]" if telefone else ""
                        print(f"  [{c}] {safe(nome_raw)}{marca}: {'[NÃO LIDA] ' if nao_lida else ''}{safe(texto[:60])}")

                    # Dedup: sem texto nao processar repetido
                    agora = time.time()
                    if not texto:
                        ultimo = self.ultimo_processamento.get(telefone, 0)
                        if agora - ultimo < 120:
                            if c % 30 == 0:
                                print(f"  [{c} SKIP] {safe(nome_raw)}: sem texto, processado há {agora-ultimo:.0f}s")
                            continue

                    # Dedup: mesma mensagem do usuario ja processada com sucesso
                    if texto:
                        dedup_window = 30 if texto.strip().isdigit() else 600
                        chave = f"{telefone}|{self._n(texto)}"
                        ult_visto = self.ultimo_visto_texto.get(chave, 0)
                        if agora - ult_visto < dedup_window:
                            if c % 30 == 0:
                                print(f"  [{c} SKIP] {safe(nome_raw)}: texto já processado ({agora-ult_visto:.0f}s atrás)")
                            continue

                    # Guarda forte: se bot enviou msg nos ultimos 8s, ignora (evita loop propria msg)
                    if texto and agora - self.ultimo_envio.get(telefone, 0) < 8:
                        if c % 30 == 0:
                            print(f"  [{c} SKIP] {safe(nome_raw)}: envio recente ({agora-self.ultimo_envio.get(telefone,0):.1f}s)")
                        continue

                    # Pula se o texto da sidebar for igual ao que o bot acabou de enviar
                    ultimo_env = self.ultimo_envio_texto.get(telefone, "")
                    if texto and ultimo_env:
                        texto_norm = re.sub(r'\s+', ' ', self._strip_emoji(texto)).strip().lower()
                        envio_norm = re.sub(r'\s+', ' ', self._strip_emoji(ultimo_env)).strip().lower()
                        if (envio_norm.startswith(texto_norm) or texto_norm.startswith(envio_norm)) and len(texto_norm) >= 20:
                            if c % 30 == 0:
                                print(f"  [{c} SKIP] {safe(nome_raw)}: texto igual ao último envio")
                            continue

                    # Fallback: detectar por mudanca de texto (msgs sem badge)
                    ultimo_texto = self.ultimo_texto_chat.get(telefone, "")
                    if texto and texto != ultimo_texto:
                        self.ultimo_texto_chat[telefone] = texto
                        if not nao_lida:
                            nao_lida = True
                    elif texto and texto == ultimo_texto and nao_lida:
                        ult_env = self.ultimo_envio.get(telefone, 0)
                        if ult_env > 0 and agora - ult_env < 30:
                            nao_lida = False
                        elif ult_env == 0 and telefone in self.chats_com_resposta:
                            if c % 30 == 0:
                                print(f"  [{c} SKIP] {safe(nome_raw)}: conversa antiga sem nova mensagem")
                            nao_lida = False

                    # Primeiro ciclo: popula ultimo_texto_chat e descobre chats com resposta previa
                    if self.primeiro_ciclo:
                        if telefone not in self.chats_com_resposta and not texto:
                            pass  # sem texto na sidebar, sem info
                        elif telefone not in self.chats_com_resposta:
                            conversa = get_conversa_ativa(telefone)
                            if conversa:
                                h = get_historico_conversa(conversa["conversa_id"], limite=1)
                                if any(m["origem"] == "agente" for m in h):
                                    self.chats_com_resposta.add(telefone)
                        continue

                    if nao_lida:
                        self.ultimo_processamento[telefone] = agora
                        print(f"\n>>> NOVA MENSAGEM: {safe(nome_raw)}: {safe(texto)}", flush=True)
                        if telefone not in self.fila_pendentes:
                            self.fila_pendentes.add(telefone)
                            await self.fila_mensagens.put((nome_key, texto, telefone, nome_raw))
                        continue

                self.primeiro_ciclo = False

                if c % 600 == 0:
                    await self._limpar_dicts_antigos()

                await asyncio.sleep(0.8)

            except asyncio.CancelledError:
                break
            except json.JSONDecodeError:
                print(f"  [DEBUG] JSON inválido: {safe(raw)[:100]}")
                erros_consecutivos += 1
                await asyncio.sleep(5)
            except Exception as e:
                msg = str(e).lower()
                if "context" in msg or "navigation" in msg or "target closed" in msg:
                    print(f"  [NAV] Página perdida, tentando recuperar...")
                    await self._recuperar_pagina()
                    erros_consecutivos = 0
                else:
                    print(f"[ERRO] {safe(e)}")
                    import traceback
                    traceback.print_exc()
                    erros_consecutivos += 1
                await asyncio.sleep(5)

            if erros_consecutivos > 10:
                print("[AVISO] Muitos erros consecutivos. Reiniciando página...")
                await self._recuperar_pagina()
                erros_consecutivos = 0

    async def _enviar_com_evaluate(self, acao_js: str, max_tentativas: int = 30):
        for _ in range(max_tentativas):
            try:
                ok = await self.avaliar(acao_js)
                if ok:
                    return True
            except Exception:
                pass
            await asyncio.sleep(0.5)
        return False

    SELETOR_INPUT = '#main [contenteditable="true"]'

    async def _aguardar_input(self, timeout=4):
        for _ in range(timeout * 2):
            el = await self.page.query_selector(self.SELETOR_INPUT)
            if el:
                return el
            await asyncio.sleep(0.25)
        return None

    async def _digitar(self, texto: str):
        caixa = await self._aguardar_input(4)
        if not caixa:
            print("  [DIG] Input não encontrado")
            return False
        try:
            await caixa.fill(texto)
            return True
        except Exception:
            try:
                await caixa.evaluate("el => { el.focus(); el.innerHTML = ''; }")
                await self.page.keyboard.type(texto, delay=20)
                return True
            except Exception as e:
                print(f"  [DIG] Falha ao digitar: {safe(str(e)[:60])}")
                return False

    async def _clicar_enviar(self, max_tentativas=15, usar_enter=False):
        for i in range(max_tentativas):
            if usar_enter:
                try:
                    await self.page.keyboard.press("Enter")
                    await asyncio.sleep(0.2)
                    ok = await self.avaliar("""
                        () => {
                            const spans = document.querySelectorAll('span[data-icon="send"]');
                            for (const sp of spans) {
                                const r = sp.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) return false;
                            }
                            return true;
                        }
                    """)
                    if ok:
                        return True
                except:
                    pass

            ok = await self.avaliar("""
                () => {
                    const vis = el => { const r = el.getBoundingClientRect(); return r.width > 0 && r.height > 0; };
                    const botoes = document.querySelectorAll('button');
                    for (const btn of botoes) {
                        const label = (btn.getAttribute('aria-label') || '').toLowerCase();
                        if ((label.includes('enviar') || label.includes('send')) && vis(btn)) {
                            btn.click();
                            return true;
                        }
                    }
                    const spans = document.querySelectorAll('span[data-icon="send"]');
                    for (const sp of spans) {
                        const pai = sp.closest('button') || sp.parentElement;
                        if (pai && vis(pai)) { pai.click(); return true; }
                        if (vis(sp)) { sp.click(); return true; }
                    }
                    const divs = document.querySelectorAll('div[role="button"]');
                    for (const d of divs) {
                        const label = (d.getAttribute('aria-label') || '').toLowerCase();
                        if ((label.includes('enviar') || label.includes('send')) && vis(d)) {
                            d.click(); return true;
                        }
                    }
                    return false;
                }
            """)
            if ok:
                return True
            await asyncio.sleep(0.25)
        return False

    async def _abrir_chat_sidebar(self, nome: str = "", telefone: str = "") -> bool:
        try:
            await self.page.wait_for_selector('#side', timeout=10000)
            if nome:
                for _ in range(5):
                    try:
                        el = self.page.get_by_title(nome).first
                        if await el.count() > 0 and await el.is_visible():
                            await el.click()
                            await asyncio.sleep(0.4)
                            panel = await self.page.query_selector('[data-testid="conversation-panel-main"]')
                            if panel:
                                return True
                    except Exception:
                        pass
                    alvo = json.dumps(nome)
                    ok = await self.avaliar(f"""
                        () => {{
                            const norm = s => s.normalize('NFKC').replace(/[\\s\\u00a0\\u200b\\u200c\\u200d\\ufeff]+/g, ' ').trim();
                            const rows = document.querySelectorAll('#side [role="row"]');
                            const alvo = {alvo};
                            const alvo_norm = norm(alvo);
                            for (const row of rows) {{
                                const el = row.querySelector('[title]');
                                if (el && norm(el.getAttribute('title')) === alvo_norm) {{
                                    row.click();
                                    return true;
                                }}
                            }}
                            return false;
                        }}
                    """)
                    if ok:
                        await asyncio.sleep(0.4)
                        return True
                    await asyncio.sleep(0.5)
            if telefone:
                mapa_json = json.dumps(self.mapa_contatos)
                for _ in range(3):
                    ok = await self.avaliar(f"""
                        (mapa) => {{
                            const tel = {json.dumps(telefone)};
                            const rows = document.querySelectorAll('#side [role="row"]');
                            for (const row of rows) {{
                                const titleEl = row.querySelector('[title]');
                                if (!titleEl) continue;
                                const name = titleEl.getAttribute('title') || '';
                                const titleTel = name.replace(/\\D/g, '');
                                // Match by digits in title (unsaved contact)
                                if (titleTel && (titleTel.endsWith(tel) || tel.endsWith(titleTel))) {{
                                    row.click();
                                    return true;
                                }}
                                // Match by name in mapa_contatos (saved contact)
                                const phoneFromMapa = mapa[name];
                                if (phoneFromMapa && (phoneFromMapa === tel || tel.endsWith(phoneFromMapa) || phoneFromMapa.endsWith(tel))) {{
                                    row.click();
                                    return true;
                                }}
                            }}
                            return false;
                        }}
                    """, self.mapa_contatos)
                    if ok:
                        await asyncio.sleep(0.4)
                        return True
                    await asyncio.sleep(0.5)
            return False
        except Exception as e:
            print(f"  [sidebar erro] {safe(str(e)[:80])}", flush=True)
            return False

    async def _chat_ja_aberto(self, nome: str) -> bool:
        if not nome:
            return False
        alvo = json.dumps(nome[:80])
        return await self.avaliar(f"""
            () => {{
                const h = document.querySelector('#main header');
                if (!h) return false;
                const el = h.querySelector('[title]');
                if (!el) return false;
                const title = (el.getAttribute('title') || '').trim();
                if (!title) return false;
                const norm = s => s.normalize('NFKC').replace(/[\\s\\u00a0\\u200b\\u200c\\u200d\\ufeff]+/g, ' ').trim();
                const alvo = {alvo};
                return norm(title).includes(norm(alvo)) || norm(alvo).includes(norm(title));
            }}
        """)

    async def enviar_texto(self, numero: str, texto: str, nome_sidebar: str = "") -> bool:
        async with self.sidebar_lock:
            try:
                if not nome_sidebar:
                    nome_sidebar = next((n for n, t in self.mapa_contatos.items() if t == numero), "")
                nomes = [nome_sidebar] if nome_sidebar else []
                nomes += [n for n, t in self.mapa_contatos.items() if t == numero]
                chat_aberto = False
                # Fast path: check if current chat is already the target
                for nome in nomes:
                    if not nome:
                        continue
                    if await self._chat_ja_aberto(nome):
                        chat_aberto = True
                        break
                if not chat_aberto:
                    for nome in nomes:
                        if not nome:
                            continue
                        if await self._abrir_chat_sidebar(nome, numero):
                            chat_aberto = True
                            break
                if not chat_aberto:
                    if await self._abrir_chat_sidebar(telefone=numero):
                        chat_aberto = True
                if not chat_aberto:
                    print(f"  -> Não foi possível abrir chat para {numero}", flush=True)
                    return False

                tem_input = await self.page.wait_for_selector(self.SELETOR_INPUT, timeout=5000)
                if not tem_input:
                    print(f"  -> Input não disponível para {numero}", flush=True)
                    return False

                ok_dig = await self._digitar(texto)
                if not ok_dig:
                    print(f"  -> Input não disponível para {numero}", flush=True)
                    return False
                agora = time.time()
                primeira_linha = texto.split("\n")[0][:80]
                antigo_envio = self.ultimo_envio.get(numero)
                antigo_texto = self.ultimo_envio_texto.get(numero)
                # Popula dicts ANTES de enviar para eliminar race com a sidebar
                self.ultimo_envio[numero] = agora
                self.ultimo_envio_texto[numero] = texto
                self.ultimo_texto_chat[numero] = primeira_linha
                if await self._clicar_enviar():
                    print(f"  -> Enviado para {numero}", flush=True)
                    self.ultimo_envio_sucesso[numero] = time.time()
                    if primeira_linha:
                        self.ultimo_visto_texto[f"{numero}|{self._n(primeira_linha)}"] = agora
                    return True

                print(f"  -> Falha ao enviar para {numero}", flush=True)
                # Rollback
                if antigo_envio is not None:
                    self.ultimo_envio[numero] = antigo_envio
                else:
                    del self.ultimo_envio[numero]
                if antigo_texto is not None:
                    self.ultimo_envio_texto[numero] = antigo_texto
                else:
                    del self.ultimo_envio_texto[numero]
                return False
            except Exception as e:
                print(f"[ERRO ENVIO] {safe(e)}", flush=True)
                import traceback
                traceback.print_exc()
                return False

    async def _clicar_anexar(self):
        btn = self.page.locator('button[aria-label="Anexar"], button[aria-label="Attach"], [data-testid="attach-file"]').first
        if await btn.count() == 0:
            return False
        try:
            await btn.wait_for(state="visible", timeout=5000)
            await btn.click()
            return True
        except:
            pass
        return False

    async def _enviar_midia_como_foto(self, caminho: str):
        await self._clicar_anexar()
        await asyncio.sleep(0.5)
        tem_pv = await self.avaliar("""
            () => {
                const el = document.querySelector('[data-testid="photo-video"]') ||
                           document.querySelector('button[aria-label="Fotos e vídeos"], button[aria-label="Photos & Videos"]');
                if (!el) return false;
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            }
        """)
        if tem_pv:
            for _ in range(2):
                try:
                    async with self.page.expect_file_chooser(timeout=5000) as fc_info:
                        await self.avaliar("""
                            () => {
                                const el = document.querySelector('[data-testid="photo-video"]') ||
                                           document.querySelector('button[aria-label="Fotos e vídeos"], button[aria-label="Photos & Videos"]');
                                if (el) el.click();
                            }
                        """)
                        await asyncio.sleep(0.3)
                    fc = await fc_info.value
                    await fc.set_files(str(caminho))
                    await asyncio.sleep(3)
                    return True
                except Exception:
                    await asyncio.sleep(1)

        # Estrategia B: input[accept*=image] diretamente
        try:
            img_inp = self.page.locator('input[accept*="image"]').first
            if await img_inp.count() > 0:
                await img_inp.set_input_files(str(caminho))
                await asyncio.sleep(3)
                return True
        except Exception:
            pass

        return False

    async def _enviar_midia_como_documento(self, caminho: str):
        tem_paperclip = await self.page.locator(
            'button[aria-label="Anexar"], button[aria-label="Attach"], [data-testid="attach-file"]'
        ).first.count() > 0
        if tem_paperclip:
            for tentativa in range(2):
                try:
                    if not await self._clicar_anexar():
                        continue
                    await asyncio.sleep(1)
                    doc = self.page.locator(
                        'button[aria-label="Documento"], button[aria-label="Document"], '
                        'button[aria-label="Documents"]'
                    ).first
                    try:
                        await doc.wait_for(state="visible", timeout=5000)
                    except:
                        continue
                    async with self.page.expect_file_chooser(timeout=10000) as fc_info:
                        await doc.click()
                        await asyncio.sleep(0.3)
                    fc = await fc_info.value
                    await fc.set_files(str(caminho))
                    await asyncio.sleep(3)
                    return True
                except Exception as e:
                    print(f"  -> Tentativa documento {tentativa+1}: {safe(str(e)[:60])}")
        # Fallback: modificar accept do input e setar arquivo diretamente
        for tentativa in range(2):
            try:
                await self.avaliar("""
                    () => {
                        const inp = document.querySelector('input[type="file"]');
                        if (inp) inp.setAttribute('accept', '*/*');
                    }
                """)
                await asyncio.sleep(0.2)
                await self.page.locator('input[type="file"]').first.set_input_files(str(caminho), timeout=5000)
                await asyncio.sleep(3)
                return True
            except:
                await asyncio.sleep(1)
        return False

    async def enviar_midia(self, numero: str, caminho: str, legenda: str = "", force_document: bool = False):
        async with self.sidebar_lock:
            try:
                chat_aberto = False
                nomes = list(dict.fromkeys(n for n, t in self.mapa_contatos.items() if t == numero))
                # Fast path: check if current chat is already the target
                for nome in nomes:
                    if not nome:
                        continue
                    if await self._chat_ja_aberto(nome):
                        chat_aberto = True
                        break
                if not chat_aberto:
                    for nome in nomes:
                        if not nome:
                            continue
                        if await self._abrir_chat_sidebar(nome, numero):
                            chat_aberto = True
                            break
                    if not chat_aberto:
                        if not await self._abrir_chat_sidebar(telefone=numero):
                            print(f"  -> Não foi possível abrir chat para mídia {numero}", flush=True)
                            return
                await asyncio.sleep(0.5)

                tem_input = await self.page.query_selector(self.SELETOR_INPUT)
                if not tem_input:
                    print(f"  -> Input não disponível para mídia {numero}", flush=True)
                    return

                is_img = Path(caminho).suffix.lower() in (".jpg", ".jpeg", ".png")
                is_video = Path(caminho).suffix.lower() in (".mp4", ".mov")
                ok = False

                if not force_document and (is_img or is_video):
                    ok = await self._enviar_midia_como_foto(caminho)
                    if not ok:
                        print("  -> Fallback: enviando como documento")
                        ok = await self._enviar_midia_como_documento(caminho)
                    if not ok:
                        print("  -> Fallback: input direto")
                        try:
                            await self.page.locator('input[type="file"]').first.set_input_files(str(caminho))
                            await asyncio.sleep(3)
                            ok = True
                        except:
                            pass
                else:
                    ok = await self._enviar_midia_como_documento(caminho)
                    if not ok:
                        print("  -> Fallback: input direto")
                        try:
                            await self.page.locator('input[type="file"]').first.set_input_files(str(caminho))
                            await asyncio.sleep(3)
                            ok = True
                        except:
                            pass

                if not ok:
                    print(f"  -> Não foi possível anexar: {Path(caminho).name}")
                    return

                if legenda:
                    cap = await self.page.query_selector('[data-testid="caption-input"]')
                    if cap:
                        try:
                            await cap.fill("")
                            await cap.type(legenda, delay=20)
                        except:
                            await cap.evaluate("el => el.focus()")
                            await self.page.keyboard.type(legenda)
                    await asyncio.sleep(0.5)

                if is_video:
                    if await self._clicar_enviar(20, usar_enter=False):
                        print(f"  -> Mídia enviada: {Path(caminho).name}")
                        nome_midia = next((n for n, t in self.mapa_contatos.items() if t == numero), None)
                        if nome_midia:
                            self.ultimo_texto_chat[self._n(nome_midia)] = f"📷 {Path(caminho).name}"
                        await asyncio.sleep(1)
                        return
                elif await self._clicar_enviar(15, usar_enter=True):
                    print(f"  -> Mídia enviada: {Path(caminho).name}")
                    nome_midia = next((n for n, t in self.mapa_contatos.items() if t == numero), None)
                    if nome_midia:
                        self.ultimo_texto_chat[self._n(nome_midia)] = f"📷 {Path(caminho).name}"
                    await asyncio.sleep(1)
                    return

                print(f"  -> Falha ao enviar mídia: {Path(caminho).name}")
            except Exception as e:
                print(f"[ERRO MÍDIA] {safe(e)}")

    async def _consultar_cep(self, cep: str) -> dict | None:
        url = f"https://viacep.com.br/ws/{cep}/json/"
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.get(url)
                if resp.status_code == 200:
                    data = resp.json()
                    if "erro" not in data:
                        return data
        except Exception:
            pass
        return None

    async def enviar_para_cliente(self, numero: str, texto: str, nome_sidebar: str = "") -> bool:
        return await self.enviar_texto(numero, texto, nome_sidebar)

    async def enviar_midia_para_cliente(self, numero: str, caminho: str, legenda: str = "", force_document: bool = False):
        await self.enviar_midia(numero, caminho, legenda, force_document)

    def _gerar_id_frete(self, telefone: str) -> str:
        return telefone

    async def solicitar_frete_transportadora(self, transportadora: dict, produto, cliente_info: dict, request_id: str):
        nome = cliente_info.get("nome", "N/I")
        msg = (
            f"SOLICITAÇÃO DE COTAÇÃO DE FRETE\n"
            f"{'='*30}\n"
            f"Cliente: {nome}\n"
            f"CPF/CNPJ: {cliente_info.get('cpf_cnpj', 'N/I')}\n"
            f"{'='*30}\n"
            f"Produto: {produto['nome']}\n"
            f"Dimensões: {produto['medidas']}  Peso: {produto['peso']}\n"
            f"Endereço: {cliente_info.get('endereco', 'N/I')} - "
            f"{cliente_info.get('cidade', 'N/I')}/{cliente_info.get('estado', 'N/I')} "
            f"CEP: {cliente_info.get('cep', 'N/I')}\n"
            f"{'='*30}\n"
            f"Favor retornar as informações abaixo:\n\n"
            f"CPF/CNPJ: {cliente_info.get('cpf_cnpj', 'N/I')}\n"
            f"Protocolo Transportadora: \n"
            f"VALOR DO FRETE: R$ \n"
            f"PRAZO DE ENTREGA:    dias úteis"
        )
        await self.enviar_texto(transportadora["numero"], msg)

    async def _enviar_folder(self, conv_id: int, telefone: str, produto: dict):
        md = BASE_DIR / "media" / "churrasqueiras" / produto["midia_dir"]
        folder = md / "folder.jpg"
        if folder.exists():
            await self.enviar_midia_para_cliente(telefone, folder, produto["nome"], force_document=True)
            salvar_mensagem(conv_id, "agente", "[folder.jpg]", "foto")
        else:
            await self.enviar_para_cliente(telefone, "Folder não disponível para este produto.")
            salvar_mensagem(conv_id, "agente", "Folder não disponível.")

    async def _enviar_foto(self, conv_id: int, telefone: str, produto: dict):
        md = BASE_DIR / "media" / "churrasqueiras" / produto["midia_dir"]
        fotos = sorted([f for f in md.glob("*") if f.suffix.lower() in (".jpg", ".jpeg", ".png")])
        if fotos:
            await self.enviar_midia_para_cliente(telefone, fotos[0], produto["nome"])
            salvar_mensagem(conv_id, "agente", f"[foto: {fotos[0].name}]", "foto")
        else:
            await self.enviar_para_cliente(telefone, "Foto não disponível para este produto.")

    async def _enviar_video(self, conv_id: int, telefone: str, produto: dict):
        md = BASE_DIR / "media" / "churrasqueiras" / produto["midia_dir"]
        videos = [f for f in md.glob("*") if f.suffix.lower() in (".mp4", ".mov")]
        if videos:
            await self.enviar_midia_para_cliente(telefone, videos[0], produto["nome"], force_document=True)
            salvar_mensagem(conv_id, "agente", f"[video: {videos[0].name}]", "video")
        else:
            await self.enviar_para_cliente(telefone, "Vídeo não disponível para este produto.")

    def _parse_endereco(self, endereco: str) -> dict:
        info = {"endereco": endereco}
        cep_match = re.search(r"(\d{5}-?\d{3})", endereco)
        if cep_match:
            info["cep"] = cep_match.group(1)
        else:
            ultimos_digitos = re.findall(r"\d+", endereco)
            for seq in reversed(ultimos_digitos):
                if len(seq) >= 5:
                    info["cep"] = seq
                    break

        ufs = {"SP","RJ","MG","RS","PR","SC","BA","DF","GO","MT","MS",
               "ES","CE","RN","PE","PB","AL","SE","PI","MA","PA","AM",
               "AC","RO","RR","AP","TO"}

        # Method 1: comma-separated (cidade/UF separados por virgula)
        partes_virgula = [p.strip() for p in endereco.replace("\n", ",").split(",") if p.strip()]
        if len(partes_virgula) >= 3:
            for i, parte in enumerate(partes_virgula):
                palavras = parte.split()
                for p in palavras:
                    if p.upper() in ufs:
                        info["estado"] = p.upper()
                        if i > 0:
                            info["cidade"] = partes_virgula[i - 1]
                        break
                if "estado" in info:
                    break

        # Method 2: space-separated fallback (sem virgulas)
        if "estado" not in info:
            partes = endereco.replace(",", " ").split()
            for i, p in enumerate(partes):
                if p.upper() in ufs:
                    info["estado"] = p.upper()
                    palavras_cidade = []
                    for j in range(i - 1, -1, -1):
                        if not any(c.isdigit() for c in partes[j]):
                            palavras_cidade.insert(0, partes[j])
                        else:
                            break
                    if palavras_cidade:
                        info["cidade"] = " ".join(palavras_cidade)
                    break
        return info

    def _limpar_endereco(self, endereco: str, cidade: str, estado: str, cep: str) -> str:
        clean = endereco
        if cep:
            clean = re.sub(rf"\s*,?\s*{re.escape(cep)}\s*", "", clean)
        if estado:
            clean = re.sub(rf"\s*,?\s*{re.escape(estado)}\s*,?\s*$", "", clean)
        if cidade:
            clean = re.sub(rf"\s*,?\s*{re.escape(cidade)}\s*,?\s*$", "", clean)
        clean = re.sub(r"\s*,\s*", ",", clean)
        clean = re.sub(r",+", ",", clean)
        return clean.strip().strip(",").strip()

    def _salvar_solicitacao_frete(self, telefone, nome, cpf_cnpj, endereco, cidade, estado, cep):
        from openpyxl import Workbook
        pasta = PASTA_SOLICITACOES
        pasta.mkdir(exist_ok=True)
        agora = datetime.now()
        nome_limpo = re.sub(r"[^\w]", "", nome)[:20] or "cliente"
        nome_arquivo = f"Solicitacao_Frete_{nome_limpo}_{agora.strftime('%d-%m-%Y')}_{agora.strftime('%H%M')}.xlsx"
        caminho = pasta / nome_arquivo
        cabecalhos = ["CAMPO", "VALOR"]
        endereco_limpo = self._limpar_endereco(endereco, cidade, estado, cep)
        conversa = get_conversa_ativa(telefone)
        produto_id = (conversa or {}).get("produto_interesse_id")
        produto = produto_por_id(produto_id) if produto_id else None
        dados = [
            ("DATA/HORA", agora.strftime("%d/%m/%Y %H:%M")),
            ("NOME", nome),
            ("WHATSAPP", telefone),
            ("CPF/CNPJ", cpf_cnpj),
            ("ENDERECO", endereco_limpo),
            ("CIDADE", cidade),
            ("ESTADO", estado),
            ("CEP", cep),
            ("PRODUTO", produto["nome"] if produto else "N/I"),
            ("VALOR DA NF", f"R$ {produto['preco']:.2f}" if produto else "N/I"),
            ("MEDIDAS DOS VOLUMES", produto["medidas"] if produto else "N/I"),
            ("EMBALAGEM", "PLÁSTICO BOLHA"),
            ("PESO", produto["peso"] if produto else "N/I"),
            ("VALOR DO FRETE FOB", ""),
            ("PROTOCOLO TRANSPORTADORA", ""),
            ("PRAZO DE ENTREGA", ""),
            ("STATUS", "Pendente"),
        ]
        wb = Workbook()
        ws = wb.active
        ws.append(cabecalhos)
        for campo, valor in dados:
            ws.append([campo, valor])
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len
        wb.save(caminho)
        print(f"  -> Solicitação salva: {nome_arquivo}", flush=True)
        return caminho

    def _atualizar_xlsx_venda_paga(self, xlsx_path: str):
        if not xlsx_path or not os.path.exists(xlsx_path):
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == "STATUS":
                    row[1].value = "Pago"
                    break
            wb.save(xlsx_path)
            print(f"  -> Status atualizado para 'Pago' em {os.path.basename(xlsx_path)}", flush=True)
        except Exception as e:
            print(f"  [xlsx] Erro ao atualizar venda paga: {safe(str(e)[:80])}", flush=True)

    async def _ler_msg_anterior_usuario(self):
        try:
            raw = await self.page.evaluate("""
                () => {
                    const painel = document.querySelector('#main [data-testid="conversation-panel-messages"]');
                    if (!painel) return '';
                    const msgs = painel.querySelectorAll(':scope .message-in');
                    const usuarios = [];
                    for (const el of msgs) {
                        const spans = el.querySelectorAll('span[dir="ltr"], span[dir="auto"]');
                        let textoCompleto = '';
                        for (const s of spans) {
                            const t = s.textContent.trim();
                            if (t) textoCompleto += t + ' ';
                        }
                        const t = textoCompleto.trim();
                        if (t) usuarios.push(t);
                    }
                    if (usuarios.length === 0) return '';
                    return usuarios[usuarios.length - 1];
                }
            """)
            return raw.strip()
        except Exception:
            return ""

    async def _ler_header_chat(self) -> str:
        raw = await self.avaliar("""
            () => {
                const h = document.querySelector('#main header');
                if (!h) return '';
                const el = h.querySelector('[title]');
                return el ? (el.getAttribute('title') || '').trim() : '';
            }
        """)
        return raw.strip()

    TRANSPORTADORA_FOB = "555199769477"

    def _telefones_transportadoras(self) -> dict[str, str]:
        trans = {t["nome"]: t["numero"] for t in TRANSPORTADORAS}
        trans["FOB"] = self.TRANSPORTADORA_FOB
        return trans

    async def _executar_frete(self, conv_id: int, cliente_id: int, telefone: str, nome_sidebar: str = ""):
        cliente = cliente_por_telefone(telefone)
        if not cliente:
            await self.enviar_para_cliente(telefone, "Erro ao recuperar seus dados.")
            return
        conversa = get_conversa_ativa(telefone)
        produto_id = (conversa or {}).get("produto_interesse_id")
        if not produto_id:
            await self.enviar_para_cliente(telefone, "Produto não identificado. Escolha um produto primeiro.")
            return
        produto = produto_por_id(produto_id)
        if not produto:
            await self.enviar_para_cliente(telefone, "Produto não encontrado.")
            return

        ci = {
            "nome": cliente.get("nome", "N/I"),
            "cpf_cnpj": cliente.get("cpf_cnpj", ""),
            "endereco": cliente.get("endereco", ""),
            "cidade": cliente.get("cidade", ""),
            "cep": cliente.get("cep", ""),
            "estado": cliente.get("estado", ""),
        }
        request_id = self._gerar_id_frete(telefone)
        transportadoras_reg = {}
        for t in TRANSPORTADORAS:
            cot_id = criar_cotacao(conv_id, t["nome"])
            transportadoras_reg[t["nome"]] = {
                "telefone": t["numero"], "cot_id": cot_id,
                "enviado_em": time.time(), "respondido": False,
                "texto_antes": "", "tentativas": 0,
            }
        transportadoras_reg["FOB"] = {
            "telefone": self.TRANSPORTADORA_FOB, "cot_id": criar_cotacao(conv_id, "FOB"),
            "enviado_em": time.time(), "respondido": False,
            "texto_antes": "", "tentativas": 0,
        }
        self.fretes_pendentes[request_id] = {
            "telefone": telefone,
            "nome_sidebar": nome_sidebar,
            "conv_id": conv_id,
            "produto": produto,
            "cliente_info": ci,
            "transportadoras": transportadoras_reg,
            "status": "enviado",
        }
        # Inicia marcador de timestamps para cada transportadora (evita reprocessar msgs antigas)
        agora = time.time()
        for t in TRANSPORTADORAS:
            self.transport_ultimo_id_msg[t["numero"]] = str(agora)
        self.transport_ultimo_id_msg[self.TRANSPORTADORA_FOB] = str(agora)
        # Captura texto_antes de cada transportadora ANTES de enviar (para ignorar msgs antigas)
        for nome, reg in list(transportadoras_reg.items()):
            async with self.sidebar_lock:
                if await self._abrir_chat_sidebar(telefone=reg["telefone"]):
                    await asyncio.sleep(0.5)
                    msg_atual = await self._ler_msg_anterior_usuario()
                    if msg_atual:
                        self.fretes_pendentes[request_id]["transportadoras"][nome]["texto_antes"] = msg_atual
        await self.enviar_para_cliente(telefone,
            f"Consultando fretes... (Tel: {request_id})")
        # Envia para transportadoras A/B via sidebar
        for t in TRANSPORTADORAS:
            await self.solicitar_frete_transportadora(t, produto, ci, request_id)
        # Envia para FOB via transport page (sem page.goto na pagina principal)
        await self._enviar_fob_msg(produto, ci, request_id)

    async def _enviar_fob_msg(self, produto, cliente_info: dict, request_id: str):
        nome = cliente_info.get("nome", "N/I")
        endereco = cliente_info.get("endereco", "N/I")
        msg = (
            f"SOLICITAÇÃO DE COTAÇÃO DE FRETE\n"
            f"{'='*30}\n"
            f"Cliente: {nome}\n"
            f"CPF/CNPJ: {cliente_info.get('cpf_cnpj', 'N/I')}\n"
            f"Produto: {produto['nome']}\n"
            f"NF: R$ {produto['preco']:.2f}\n"
            f"Medidas: {produto['medidas']}  Peso: {produto['peso']}\n"
            f"Endereço: {endereco}\n"
            f"{'='*30}\n"
            f"Favor retornar as informações abaixo:\n\n"
            f"CPF/CNPJ: {cliente_info.get('cpf_cnpj', 'N/I')}\n"
            f"Protocolo Transportadora: \n"
            f"VALOR DO FRETE: R$ \n"
            f"PRAZO DE ENTREGA:    dias úteis"
        )
        tp = self.transport_page
        if not tp:
            print(f"  [frete] Sem transport_page, pulando FOB", flush=True)
            return
        # Tenta enviar via sidebar na transport_page
        ok = await self._transport_enviar_texto(tp, self.TRANSPORTADORA_FOB, msg)
        if ok:
            self.ultimo_envio[self.TRANSPORTADORA_FOB] = time.time()
            self.ultimo_envio_texto[self.TRANSPORTADORA_FOB] = msg
            print(f"  -> FOB enviado via transport page sidebar (Tel: {request_id})", flush=True)
            if request_id in self.fretes_pendentes:
                for t in self.fretes_pendentes[request_id]["transportadoras"].values():
                    if t["telefone"] == self.TRANSPORTADORA_FOB:
                        t["_inicializado"] = True
            return
        # Fallback: navegar transport_page para FOB (nao afeta pagina principal)
        print(f"  [frete] FOB não encontrado na sidebar da transport_page, navegando direto...", flush=True)
        url_fob = f"https://web.whatsapp.com/send/?phone={self.TRANSPORTADORA_FOB}"
        try:
            await tp.goto(url_fob, timeout=30000)
            await asyncio.sleep(2)
            caixa = await self._transport_aguardar_input(tp, timeout=8)
            if not caixa:
                print(f"  [frete] Input não encontrado na transport_page", flush=True)
                # Voltar transport page para whatsapp principal
                await tp.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=20000)
                return
            try:
                await caixa.fill(msg)
            except:
                await caixa.evaluate("el => { el.focus(); el.innerHTML = ''; }")
                await tp.keyboard.type(msg, delay=20)
            await tp.keyboard.press("Enter")
            await asyncio.sleep(1)
            self.ultimo_envio[self.TRANSPORTADORA_FOB] = time.time()
            self.ultimo_envio_texto[self.TRANSPORTADORA_FOB] = msg
            print(f"  -> FOB enviado via transport_page goto (Tel: {request_id})", flush=True)
            if request_id in self.fretes_pendentes:
                for t in self.fretes_pendentes[request_id]["transportadoras"].values():
                    if t["telefone"] == self.TRANSPORTADORA_FOB:
                        t["_inicializado"] = True
            # Voltar transport page para whatsapp principal
            await tp.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=20000)
            await asyncio.sleep(1)
        except Exception as e:
            print(f"  [frete] Erro ao enviar para FOB: {safe(str(e)[:60])}", flush=True)

    async def _finalizar_coleta_frete(self, conv_id: int, cliente_id: int, telefone: str, nome_sidebar: str = ""):
        cliente_completo = cliente_por_telefone(telefone)
        nome_cliente = cliente_completo.get("nome", "") if cliente_completo else ""
        endereco = cliente_completo.get("endereco", "") if cliente_completo else ""
        cidade = cliente_completo.get("cidade", "") if cliente_completo else ""
        estado = cliente_completo.get("estado", "") if cliente_completo else ""
        cep = cliente_completo.get("cep", "") if cliente_completo else ""
        print(f"  [frete] endereço completo -> enviando confirmação + xlsx", flush=True)
        caminho_xlsx = self._salvar_solicitacao_frete(telefone, nome_cliente,
            cliente_completo.get("cpf_cnpj", "") if cliente_completo else "",
            endereco, cidade, estado, cep)
        await self.enviar_para_cliente(telefone,
            f"Obrigado, {nome_cliente}! Sua solicitação de frete foi recebida com sucesso.\n"
            f"Estou consultando a transportadora, aguarde um momento...")
        await self._solicitar_frete_fob(conv_id, telefone, nome_sidebar, str(caminho_xlsx))

    async def _solicitar_frete_fob(self, conv_id: int, telefone: str, nome_sidebar: str = "", xlsx_path: str = ""):
        cliente = cliente_por_telefone(telefone)
        if not cliente:
            await self.enviar_para_cliente(telefone, "Erro ao recuperar seus dados.")
            return
        conversa = get_conversa_ativa(telefone)
        produto_id = (conversa or {}).get("produto_interesse_id")
        if not produto_id:
            await self.enviar_para_cliente(telefone, "Produto não identificado.")
            return
        produto = produto_por_id(produto_id)
        if not produto:
            await self.enviar_para_cliente(telefone, "Produto não encontrado.")
            return

        ci = {
            "nome": cliente.get("nome", "N/I"),
            "cpf_cnpj": cliente.get("cpf_cnpj", ""),
            "endereco": cliente.get("endereco", "N/I"),
            "cidade": cliente.get("cidade", "N/I"),
            "estado": cliente.get("estado", "N/I"),
            "cep": cliente.get("cep", "N/I"),
        }
        request_id = self._gerar_id_frete(telefone)
        cot_id = criar_cotacao(conv_id, "FOB")
        self.fretes_pendentes[request_id] = {
            "telefone": telefone,
            "nome_sidebar": nome_sidebar,
            "conv_id": conv_id,
            "xlsx_path": xlsx_path,
            "produto": produto,
            "cliente_info": ci,
            "transportadoras": {
                "FOB": {
                    "telefone": self.TRANSPORTADORA_FOB, "cot_id": cot_id,
                    "enviado_em": time.time(), "respondido": False,
                    "texto_antes": "", "tentativas": 0,
                },
            },
            "status": "enviado",
        }
        self.transport_ultimo_id_msg[self.TRANSPORTADORA_FOB] = str(time.time())
        # Persiste xlsx_path na cotacao no banco
        if xlsx_path:
            from database import get_connection
            conn = get_connection()
            conn.execute("UPDATE cotacoes SET xlsx_path=? WHERE id=?", (str(xlsx_path), cot_id))
            conn.commit()
            conn.close()
        await self.enviar_para_cliente(telefone,
            f"Consultando frete FOB... (Tel: {request_id})")
        await self._enviar_fob_msg(produto, ci, request_id)

    async def _ler_sidebar_text_por_nome(self, nome: str) -> str:
        nome_json = json.dumps(nome)
        raw = await self.avaliar(f"""
            () => {{
                const rows = document.querySelectorAll('#side [role="row"]');
                for (const row of rows) {{
                    const el = row.querySelector('[title]');
                    if (!el) continue;
                    if (el.getAttribute('title') === {nome_json}) {{
                        const spans = row.querySelectorAll('span[dir]');
                        const titulo = el.getAttribute('title');
                        for (const sp of spans) {{
                            if (sp.getAttribute('title') !== titulo && sp.textContent.trim()) {{
                                return sp.textContent.trim();
                            }}
                        }}
                    }}
                }}
                return '';
            }}
        """)
        return raw.strip()

    async def _chat_tem_badge_nao_lida(self, nome: str) -> bool:
        nome_json = json.dumps(nome)
        raw = await self.avaliar(f"""
            () => {{
                const rows = document.querySelectorAll('#side [role="row"]');
                for (const row of rows) {{
                    const el = row.querySelector('[title]');
                    if (!el) continue;
                    if (el.getAttribute('title') === {nome_json}) {{
                        const badge = row.querySelector('[data-testid="icon-unread-count"]')
                            || row.querySelector('[aria-label*="nao lida"]')
                            || row.querySelector('[aria-label*="unread"]');
                        return !!badge;
                    }}
                }}
                return false;
            }}
        """)
        return bool(raw)

    async def _transport_processar_fretes(self):
        if not self.fretes_pendentes:
            return
        tp = self.transport_page
        if not tp:
            print("  [transport] Sem transport_page, recriando...", flush=True)
            await self._iniciar_transport_page()
            tp = self.transport_page
            if not tp:
                return
        try:
            novas = await self._transport_detecta_novas_mensagens(tp)
        except Exception as e:
            print(f"  [transport] Erro detecção: {safe(str(e)[:60])}", flush=True)
            return
        if not novas:
            return
        for tel_trans, msg_text in novas:
            await self._transport_processar_resposta(tp, tel_trans, msg_text)

    async def _transport_detecta_novas_mensagens(self, tp):
        transport_tels = [t["numero"] for t in TRANSPORTADORAS] + [self.TRANSPORTADORA_FOB]
        codigo = f"""
            async (selfNum) => {{
                const db = await new Promise(r => {{
                    const req = indexedDB.open('model-storage');
                    req.onsuccess = () => r(req.result);
                }});
                const tels = {json.dumps(transport_tels)};
                const out = [];
                const tx = db.transaction('message', 'readonly');
                const store = tx.objectStore('message');
                const all = await new Promise(r => {{
                    const req = store.getAll();
                    req.onsuccess = () => r(req.result);
                }});
                for (const m of all) {{
                    if (!m.id) continue;
                    let fromUser = '';
                    if (m.from && typeof m.from === 'string') {{
                        const match = m.from.match(/^(\\d+)@/);
                        if (match && match[1] !== selfNum) fromUser = match[1];
                    }}
                    if (!fromUser && m.to) {{
                        const toUser = typeof m.to === 'object' ? (m.to.user || '') :
                                       (typeof m.to === 'string' ? m.to.split('@')[0] : '');
                        if (/^\\d+$/.test(toUser) && toUser !== selfNum) fromUser = toUser;
                    }}
                    if (!fromUser || !tels.includes(fromUser)) continue;
                    let texto = '';
                    if (typeof m.body === 'string') texto = m.body;
                    if (m.message && m.message.conversation) texto = m.message.conversation;
                    const ts = m.t || m.messageTimestamp || 0;
                    out.push({{tel: fromUser, id: m.id, texto, ts}});
                }}
                const last = {{}};
                for (const item of out) {{
                    const prev = last[item.tel];
                    if (!prev || item.ts > prev.ts) last[item.tel] = item;
                }}
                return JSON.stringify(Object.values(last));
            }}
        """
        try:
            raw = await tp.evaluate(codigo, SEU_NUMERO)
        except Exception as e:
            print(f"  [transport] Erro IndexedDB: {safe(str(e)[:60])}", flush=True)
            return []
        try:
            mensagens = json.loads(raw)
        except json.JSONDecodeError:
            return []
        novas = []
        for m in mensagens:
            tel = m["tel"]
            texto = m.get("texto", "")
            ts = float(m.get("ts", 0))
            if not texto or not ts:
                continue
            ult_ts = float(self.transport_ultimo_id_msg.get(tel, 0))
            if ts > ult_ts:
                self.transport_ultimo_id_msg[tel] = str(ts)
                novas.append((tel, texto))
        return novas

    async def _transport_processar_resposta(self, tp, tel_trans, texto_resposta):
        req_id = None
        req = None
        for rid, r in list(self.fretes_pendentes.items()):
            for trans_nome, reg in list(r["transportadoras"].items()):
                if reg["telefone"] == tel_trans and not reg["respondido"]:
                    req_id = rid
                    req = r
                    break
            if req_id:
                break
        if not req_id or not req:
            print(f"  [transport] Resposta de {tel_trans} sem frete pendente", flush=True)
            return
        trans_nome = None
        trans_reg = None
        for tn, reg in req["transportadoras"].items():
            if reg["telefone"] == tel_trans:
                trans_nome = tn
                trans_reg = reg
                break
        if not trans_nome or not trans_reg:
            return
        print(f"  [transport] Resposta detectada de {trans_nome} para Tel {req_id}", flush=True)
        valor = self.extrair_valor_frete(texto_resposta)
        prazo = self.extrair_prazo(texto_resposta)
        prot_transp = self.extrair_protocolo_transportadora(texto_resposta)
        print(f"  [transport] Extraido -> R$ {valor:.2f}, prazo={prazo or 'None'}, prot={prot_transp or 'None'}", flush=True)
        msg_cliente = f"Frete {trans_nome} (Tel: {req_id}):\n"
        if valor > 0:
            msg_cliente += f"Valor: R$ {valor:.2f}\n"
        else:
            msg_cliente += f"Valor: {texto_resposta}\n"
        msg_cliente += f"Prazo: {prazo or 'a confirmar'}\n"
        if prot_transp:
            msg_cliente += f"Protocolo Transportadora: {prot_transp}\n"
        if valor > 0:
            msg_cliente += f"Total c/ produto: R$ {req['produto']['preco'] + valor:.2f}\n"
        msg_cliente += f"\nDeseja confirmar o pedido?"
        tel_cliente = req["telefone"]
        ok = await self._transport_enviar_texto(tp, tel_cliente, msg_cliente, req.get("nome_sidebar", ""))
        if ok:
            trans_reg["respondido"] = True
            if trans_reg.get("cot_id"):
                atualizar_cotacao(trans_reg["cot_id"], valor_frete=valor, prazo=prazo, status="recebida")
            atualizar_etapa_conversa(req["conv_id"], "frete_confirmar")
            xlsx_path = req.get("xlsx_path")
            if xlsx_path and os.path.exists(xlsx_path):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(xlsx_path)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, max_col=2):
                        campo = row[0].value
                        if campo == "STATUS":
                            row[1].value = "Cotado"
                        elif campo == "VALOR DO FRETE FOB" and valor > 0:
                            row[1].value = f"R$ {valor:.2f}"
                        elif campo == "PROTOCOLO TRANSPORTADORA" and prot_transp:
                            row[1].value = prot_transp
                        elif campo == "PRAZO DE ENTREGA" and prazo:
                            row[1].value = prazo
                    wb.save(xlsx_path)
                    print(f"  -> Status atualizado em {os.path.basename(xlsx_path)}", flush=True)
                except Exception as e2:
                    print(f"  [transport] Erro xlsx: {safe(str(e2)[:60])}", flush=True)
            print(f"  [transport] Resposta {trans_nome} encaminhada ao cliente", flush=True)
        else:
            print(f"  [transport] Falha ao enviar para cliente {tel_cliente}, tentará novamente", flush=True)
        if all(t["respondido"] for t in req["transportadoras"].values()):
            self.fretes_pendentes.pop(req_id, None)

    async def _transport_enviar_texto(self, tp, numero, texto, nome_sidebar=""):
        try:
            if not nome_sidebar:
                nome_sidebar = next((n for n, t in self.mapa_contatos.items() if t == numero), "")
            nomes = [nome_sidebar] if nome_sidebar else []
            nomes += [n for n, t in self.mapa_contatos.items() if t == numero]
            chat_aberto = False
            for nome in nomes:
                if not nome:
                    continue
                if await self._transport_chat_ja_aberto(tp, nome):
                    chat_aberto = True
                    break
            if not chat_aberto:
                for nome in nomes:
                    if not nome:
                        continue
                    if await self._transport_abrir_chat_sidebar(tp, nome, numero):
                        chat_aberto = True
                        break
            if not chat_aberto:
                if await self._transport_abrir_chat_sidebar(tp, telefone=numero):
                    chat_aberto = True
            if not chat_aberto:
                print(f"  [transport] Não foi possível abrir chat para {numero}", flush=True)
                return False
            caixa = await self._transport_aguardar_input(tp)
            if not caixa:
                print(f"  [transport] Input não disponível para {numero}", flush=True)
                return False
            try:
                await caixa.fill(texto)
            except Exception:
                await caixa.evaluate("el => { el.focus(); el.innerHTML = ''; }")
                await tp.keyboard.type(texto, delay=20)
            await tp.keyboard.press("Enter")
            await asyncio.sleep(0.5)
            print(f"  [transport] Enviado para {numero}", flush=True)
            return True
        except Exception as e:
            print(f"  [transport] Erro envio: {safe(str(e)[:80])}", flush=True)
            return False

    async def _transport_chat_ja_aberto(self, tp, nome):
        if not nome:
            return False
        try:
            return await tp.evaluate(f"""
                () => {{
                    const h = document.querySelector('#main header');
                    if (!h) return false;
                    const el = h.querySelector('[title]');
                    if (!el) return false;
                    const title = (el.getAttribute('title') || '').trim();
                    if (!title) return false;
                    const norm = s => s.normalize('NFKC').replace(/[\\s\\u00a0\\u200b\\u200c\\u200d\\ufeff]+/g, ' ').trim();
                    const alvo = {json.dumps(nome[:80])};
                    return norm(title).includes(norm(alvo)) || norm(alvo).includes(norm(title));
                }}
            """)
        except:
            return False

    async def _transport_abrir_chat_sidebar(self, tp, nome="", telefone=""):
        try:
            await tp.wait_for_selector('#side', timeout=10000)
            if nome:
                for _ in range(5):
                    try:
                        el = tp.get_by_title(nome).first
                        if await el.count() > 0 and await el.is_visible():
                            await el.click()
                            await asyncio.sleep(0.4)
                            panel = await tp.query_selector('[data-testid="conversation-panel-main"]')
                            if panel:
                                return True
                    except:
                        pass
                    ok = await tp.evaluate(f"""
                        () => {{
                            const norm = s => s.normalize('NFKC').replace(/[\\s\\u00a0\\u200b\\u200c\\u200d\\ufeff]+/g, ' ').trim();
                            const rows = document.querySelectorAll('#side [role="row"]');
                            const alvo = {json.dumps(nome)};
                            const alvo_norm = norm(alvo);
                            for (const row of rows) {{
                                const el = row.querySelector('[title]');
                                if (el && norm(el.getAttribute('title')) === alvo_norm) {{
                                    row.click();
                                    return true;
                                }}
                            }}
                            return false;
                        }}
                    """)
                    if ok:
                        await asyncio.sleep(0.4)
                        return True
                    await asyncio.sleep(0.5)
            if telefone:
                for _ in range(3):
                    ok = await tp.evaluate(f"""
                        (mapa) => {{
                            const tel = {json.dumps(telefone)};
                            const rows = document.querySelectorAll('#side [role="row"]');
                            for (const row of rows) {{
                                const titleEl = row.querySelector('[title]');
                                if (!titleEl) continue;
                                const name = titleEl.getAttribute('title') || '';
                                const titleTel = name.replace(/\\D/g, '');
                                if (titleTel && (titleTel.endsWith(tel) || tel.endsWith(titleTel))) {{
                                    row.click();
                                    return true;
                                }}
                                const phoneFromMapa = mapa[name];
                                if (phoneFromMapa && (phoneFromMapa === tel || tel.endsWith(phoneFromMapa) || phoneFromMapa.endsWith(tel))) {{
                                    row.click();
                                    return true;
                                }}
                            }}
                            return false;
                        }}
                    """, self.mapa_contatos)
                    if ok:
                        await asyncio.sleep(0.4)
                        return True
                    await asyncio.sleep(0.5)
            return False
        except:
            return False

    async def _transport_aguardar_input(self, tp, timeout=4):
        for _ in range(timeout * 2):
            el = await tp.query_selector('#main [contenteditable="true"]')
            if el:
                return el
            await asyncio.sleep(0.25)
        return None

    async def processar_mensagem(self, remetente: str, msg_texto: str, telefone: str = "", nome_sidebar: str = ""):
        try:
            if not telefone:
                telefone = re.sub(r'\D', '', remetente)
                if not telefone.startswith("55"):
                    telefone = "55" + telefone
                if len(telefone) < 12:
                    telefone = "55" + re.sub(r'\D', '', remetente)

            if len(telefone) < 12:
                print(f"  -> Telefone inválido p/ {safe(remetente)}: {telefone}", flush=True)
                return

            # BOT-DEDUP: se o bot enviou msg nos ultimos 10s e o conteudo parece eco do bot, ignora
            if telefone and time.time() - self.ultimo_envio_sucesso.get(telefone, 0) < 10:
                ultimo_texto = self.ultimo_envio_texto.get(telefone, "")
                if ultimo_texto and msg_texto:
                    msg_norm = re.sub(r'\s+', ' ', self._strip_emoji(msg_texto)).strip().lower()
                    env_norm = re.sub(r'\s+', ' ', self._strip_emoji(ultimo_texto)).strip().lower()
                    if env_norm.startswith(msg_norm) or (len(msg_norm) > 20 and msg_norm in env_norm):
                        print(f"  [BOT-DEDUP] {safe(remetente)}: proprio bot ignorado", flush=True)
                        return
                elif not msg_texto:
                    return

            if telefone in self.processando:
                return
            self.processando[telefone] = True

            # Ignora mensagens de sistema do WhatsApp
            # Comando do vendedor: confirmar pagamento manual
            if telefone == re.sub(r"\D", "", SEU_NUMERO) and msg_texto:
                m_pagar = re.match(r"^pagar\s+(\d+)$", msg_texto.strip().lower())
                if m_pagar:
                    venda_id = int(m_pagar.group(1))
                    venda = confirmar_pagamento(venda_id)
                    if venda:
                        tel_cliente = venda.get("cliente_telefone", "")
                        nome_cliente = venda.get("cliente_nome", "")
                        print(f"  [PAGAMENTO] Venda {venda_id} confirmada manualmente para {nome_cliente}", flush=True)
                        await self.enviar_para_cliente(telefone,
                            f"✅ Pagamento confirmado! Venda #{venda_id} - {nome_cliente}")
                        if tel_cliente:
                            await self.enviar_para_cliente(tel_cliente,
                                f"✅ Pagamento confirmado! Seu pedido será processado em breve. Obrigado!")
                    else:
                        await self.enviar_para_cliente(telefone,
                            f"Venda #{venda_id} não encontrada.")
                    self.processando.pop(telefone, None)
                    return

            if msg_texto and ("Meta" in msg_texto or "servi" in msg_texto.lower() or "gerenciar esta conversa" in msg_texto.lower()):
                print(f"  -> Msg de sistema ignorada: {safe(msg_texto[:60])}", flush=True)
                self.processando.pop(telefone, None)
                return

            cliente_id = criar_cliente(telefone, nome=remetente)
            conversa = get_conversa_ativa(telefone)
            if not conversa:
                if tem_compra_finalizada(telefone):
                    conv_id = criar_conversa(cliente_id, etapa_inicial="pos_compra")
                else:
                    conv_id = criar_conversa(cliente_id)
            else:
                conv_id = conversa["conversa_id"]

            if msg_texto:
                salvar_mensagem(conv_id, "cliente", msg_texto)
            historico = get_historico_conversa(conv_id, limite=30)

            # So inicia apresentacao se ainda nao houver resposta do bot
            tem_resposta = any(m["origem"] == "agente" for m in historico)
            if not tem_resposta:
                if not conversa:
                    from database import get_connection
                    conn2 = get_connection()
                    row2 = conn2.execute("SELECT etapa FROM conversas WHERE id=?", (conv_id,)).fetchone()
                    etapa_conv = row2["etapa"] if row2 else ""
                    conn2.close()
                    if etapa_conv == "pos_compra":
                        await self._perguntar_nova_compra_ou_info(telefone, conv_id, remetente)
                        return
                ok = await self._iniciar_apresentacao_menu(telefone, conv_id, nome_sidebar)
                self.processando.pop(telefone, None)
                if ok:
                    print(f"  -> Apresentação iniciada para {safe(remetente)}", flush=True)
                else:
                    print(f"  -> Falha ao enviar menu para {safe(remetente)}", flush=True)
                return

            # Msg sem texto detectável: reinicia apresentação
            if not msg_texto:
                print(f"  -> Msg sem texto, reiniciando menu para {safe(remetente)}", flush=True)
                ok = await self._iniciar_apresentacao_menu(telefone, conv_id, nome_sidebar)
                self.processando.pop(telefone, None)
                if ok:
                    print(f"  -> Apresentação reiniciada para {safe(remetente)}", flush=True)
                else:
                    print(f"  -> Falha ao reiniciar menu para {safe(remetente)}", flush=True)
                return

            # Agradecimento: nao abre menu, apenas responde e encerra
            agradecimentos = {"obrigado", "obrigada", "valeu", "brigado", "brigada",
                              "muito obrigado", "muito obrigada", "obrigadão",
                              "agradecido", "thanks", "thank you"}
            msg_clean = msg_texto.strip().lower().rstrip("?!.")
            if msg_clean in agradecimentos:
                await self.enviar_para_cliente(telefone,
                    "😊 Por nada! Estou aqui para ajudar. É só me chamar quando precisar.",
                    nome_sidebar)
                salvar_mensagem(conv_id, "agente", "Por nada!")
                self.processando.pop(telefone, None)
                return

            etapa = (conversa or {}).get("etapa", "")

            # --- CLIENTE COM COMPRA ANTERIOR: pergunta nova compra ou informações ---
            if ((etapa in ("menu_principal", "saudacao", "") and tem_compra_finalizada(telefone)) or etapa == "fechada"):
                nome_cliente = (conversa or {}).get("nome", remetente)
                await self._perguntar_nova_compra_ou_info(telefone, conv_id, nome_cliente, alterar_etapa=True)
                return

            # --- POS-COMPRA: nova compra ou informações ---
            if etapa == "pos_compra":
                opt_pos = msg_texto.strip().lower()
                palavras_nova = ["1", "nova", "comprar", "novamente", "quero", "sim"]
                palavras_info = ["2", "info", "informações", "informacoes", "anterior", "compra anterior"]
                if any(p in opt_pos for p in palavras_nova):
                    from database import get_connection
                    conn = get_connection()
                    conn.execute("UPDATE conversas SET etapa='menu_principal' WHERE id=?", (conv_id,))
                    conn.commit()
                    conn.close()
                    print(f"  -> Conversa {conv_id} reaberta para nova compra ({safe(remetente)})", flush=True)
                    ok = await self._iniciar_apresentacao_menu(telefone, conv_id, nome_sidebar)
                    self.processando.pop(telefone, None)
                    if ok:
                        print(f"  -> Nova apresentação iniciada para {safe(remetente)}", flush=True)
                    return
                elif any(p in opt_pos for p in palavras_info):
                    await self.enviar_para_cliente(telefone,
                        "Informaremos quando for entregue à transportadora.")
                    salvar_mensagem(conv_id, "agente", "Informações da compra anterior")
                    atualizar_etapa_conversa(conv_id, "fechada")
                    self.processando.pop(telefone, None)
                    return
                else:
                    nome_cliente = (conversa or {}).get("nome", remetente)
                    await self._perguntar_nova_compra_ou_info(telefone, conv_id, nome_cliente)
                    return

            # --- FLUXO DE FRETE: coleta de dados ---
            if etapa in ("frete_nome", "frete_cpf", "frete_cep", "frete_numero"):
                cli = cliente_por_telefone(telefone)
                if cli and all([cli.get("nome"), cli.get("cpf_cnpj"), cli.get("endereco"), cli.get("cidade"), cli.get("estado"), cli.get("cep")]):
                    await self._finalizar_coleta_frete(conv_id, cli["id"], telefone, nome_sidebar)
                    return
            if etapa == "frete_nome":
                nome = msg_texto.strip()
                atualizar_cliente(cliente_id, nome=nome)
                atualizar_etapa_conversa(conv_id, "frete_cpf")
                print(f"  [frete] nome salvo: {safe(nome)} -> etapa frete_cpf", flush=True)
                await self.enviar_para_cliente(telefone, "Obrigado! Agora informe seu CPF ou CNPJ:")
                salvar_mensagem(conv_id, "agente", "Obrigado! Agora informe seu CPF ou CNPJ:")
                return

            if etapa == "frete_cpf":
                cpf_cnpj_raw = msg_texto.strip()
                digitos = re.sub(r"\D", "", cpf_cnpj_raw)
                if not (11 <= len(digitos) <= 14):
                    await self.enviar_para_cliente(telefone,
                        "CPF ou CNPJ inválido. Digite apenas números (11 dígitos para CPF, 14 para CNPJ):")
                    salvar_mensagem(conv_id, "agente", "CPF ou CNPJ inválido.")
                    return
                cpf_cnpj = digitos
                atualizar_cliente(cliente_id, cpf_cnpj=cpf_cnpj)
                atualizar_etapa_conversa(conv_id, "frete_cep")
                print(f"  [frete] cpf_cnpj salvo: {safe(cpf_cnpj)} -> etapa frete_cep", flush=True)
                await self.enviar_para_cliente(telefone, "Perfeito! Agora informe seu CEP:")
                salvar_mensagem(conv_id, "agente", "Informe o CEP:")
                return

            if etapa == "frete_cep":
                cep_raw = msg_texto.strip()
                digitos_cep = re.sub(r"\D", "", cep_raw)
                if len(digitos_cep) != 8:
                    await self.enviar_para_cliente(telefone,
                        "CEP inválido. Digite 8 dígitos:")
                    return
                dados = await self._consultar_cep(digitos_cep)
                if not dados:
                    await self.enviar_para_cliente(telefone,
                        "CEP não encontrado. Digite um CEP válido:")
                    return
                self._cache_cep[telefone] = dados
                atualizar_etapa_conversa(conv_id, "frete_numero")
                logr = dados.get("logradouro", "")
                bairro = dados.get("bairro", "")
                localidade = dados.get("localidade", "")
                uf = dados.get("uf", "")
                if logr:
                    msg = f"{logr}"
                    if bairro:
                        msg += f", {bairro}"
                    msg += f", {localidade}/{uf}\n\nQual o número do local?"
                else:
                    msg = f"{localidade}/{uf}"
                    if bairro:
                        msg = f"Bairro {bairro}, " + msg
                    msg += "\n\nQual seu endereço completo (rua e número)?"
                await self.enviar_para_cliente(telefone, msg)
                salvar_mensagem(conv_id, "agente", "Informe o número:")
                return

            if etapa == "frete_numero":
                dados_cep = self._cache_cep.pop(telefone, None)
                if not dados_cep:
                    atualizar_etapa_conversa(conv_id, "frete_cep")
                    await self.enviar_para_cliente(telefone, "Erro. Informe seu CEP novamente:")
                    return
                numero_raw = msg_texto.strip()
                logradouro = dados_cep.get("logradouro", "")
                if logradouro:
                    # ViaCEP tem rua — pede so o numero
                    if not numero_raw.isdigit():
                        self._cache_cep[telefone] = dados_cep
                        await self.enviar_para_cliente(telefone, "Digite apenas o número do local:")
                        return
                    numero = numero_raw
                    bairro = dados_cep.get("bairro", "")
                    endereco = f"{logradouro}, {numero}"
                    if bairro:
                        endereco += f" - {bairro}"
                else:
                    # Sem rua (CEP generico) — trata como endereco completo
                    endereco = numero_raw
                cidade = dados_cep.get("localidade", "")
                estado = dados_cep.get("uf", "")
                cep = dados_cep.get("cep", "").replace("-", "")
                atualizar_cliente(cliente_id, endereco=endereco, cidade=cidade, estado=estado, cep=cep)
                self._cache_endereco[telefone] = {"endereco": endereco, "cidade": cidade, "estado": estado, "cep": cep}
                atualizar_etapa_conversa(conv_id, "frete_complemento")
                print(f"  [frete] número salvo -> perguntando complemento", flush=True)
                await self.enviar_para_cliente(telefone,
                    "Deseja inserir algum complemento?\n[h] SIM - Sim\n[i] NÃO - Não")
                salvar_mensagem(conv_id, "agente", "Deseja complemento?")
                return

            # --- FRETE: perguntar se quer complemento ---
            if etapa == "frete_complemento":
                opt = self._n(msg_texto.strip().lower())
                if opt in ("h", "sim", "s"):
                    atualizar_etapa_conversa(conv_id, "frete_complemento_texto")
                    await self.enviar_para_cliente(telefone,
                        "Qual o complemento? (ex: apto 42, bloco B)")
                    salvar_mensagem(conv_id, "agente", "Informe o complemento:")
                    self.processando.pop(telefone, None)
                    return
                # Qualquer outra resposta = sem complemento
                dados_end = self._cache_endereco.pop(telefone, None)
                if dados_end:
                    atualizar_cliente(cliente_id, endereco=dados_end["endereco"],
                                     cidade=dados_end["cidade"], estado=dados_end["estado"], cep=dados_end["cep"])
                atualizar_etapa_conversa(conv_id, "frete_aguardando")
                self.processando.pop(telefone, None)
                await self._finalizar_coleta_frete(conv_id, cliente_id, telefone, nome_sidebar)
                return

            # --- FRETE: capturar texto do complemento ---
            if etapa == "frete_complemento_texto":
                complemento = msg_texto.strip()
                dados_end = self._cache_endereco.pop(telefone, None)
                if dados_end:
                    endereco_completo = f"{dados_end['endereco']} - {complemento}"
                    atualizar_cliente(cliente_id, endereco=endereco_completo,
                                     cidade=dados_end["cidade"], estado=dados_end["estado"], cep=dados_end["cep"])
                atualizar_etapa_conversa(conv_id, "frete_aguardando")
                self.processando.pop(telefone, None)
                await self._finalizar_coleta_frete(conv_id, cliente_id, telefone, nome_sidebar)
                return

            # --- FRETE: aguardando resposta da transportadora ---
            if etapa == "frete_aguardando":
                # Só bloqueia se ainda houver frete pendente para este telefone
                tem_pendente = any(
                    req["telefone"] == telefone
                    for req in self.fretes_pendentes.values()
                    if req["status"] == "enviado" and not all(t["respondido"] for t in req["transportadoras"].values())
                )
                if not tem_pendente:
                    atualizar_etapa_conversa(conv_id, "menu_principal")
                    etapa = "menu_principal"
                else:
                    await self.enviar_para_cliente(telefone,
                        "Ainda estou aguardando a resposta da transportadora. Assim que receber, aviso você!")
                    self.processando.pop(telefone, None)
                    return

            # --- FRETE: aguardando pagamento do cliente ---
            if etapa == "frete_aguardando_pagamento":
                venda_pend = get_venda_pendente_conversa(conv_id)
                if venda_pend:
                    if venda_pend.get("payment_status") == "pago":
                        atualizar_etapa_conversa(conv_id, "fechada")
                        await self.enviar_para_cliente(telefone,
                            "✅ Pagamento confirmado! Seu pedido será processado em breve.")
                        self.processando.pop(telefone, None)
                        return
                    # Verifica status no Stripe
                    sess_id = venda_pend.get("stripe_session_id")
                    if not sess_id and venda_pend.get("payment_url"):
                        m = re.search(r"(cs_test_[a-zA-Z0-9]+)", venda_pend["payment_url"])
                        if m:
                            sess_id = m.group(1)
                    if sess_id and verificar_pagamento(sess_id):
                        from database import get_connection
                        conn = get_connection()
                        conn.execute("UPDATE vendas SET payment_status='pago', status='pago' WHERE id=?", (venda_pend["id"],))
                        conn.execute("UPDATE conversas SET etapa='fechada' WHERE id=?", (conv_id,))
                        conn.commit()
                        conn.close()
                        atualizar_etapa_conversa(conv_id, "fechada")
                        await self.enviar_para_cliente(telefone,
                            "✅ Pagamento confirmado! Seu pedido será processado em breve. Obrigado!")
                        await self.enviar_para_cliente(SEU_NUMERO,
                            f"✅ PAGAMENTO CONFIRMADO - Venda #{venda_pend['id']} - {venda_pend.get('cliente_nome', '')}")
                        print(f"[PAGAMENTO] Stripe confirmou pagamento da venda {venda_pend['id']}", flush=True)
                        self.processando.pop(telefone, None)
                        return
                link = venda_pend.get("payment_url", "") if venda_pend else ""
                msg = "Seu pedido está aguardando a confirmação do pagamento.\n"
                if link:
                    msg += f"💳 Link para pagamento: {link}"
                else:
                    msg += "Assim que o pagamento for confirmado, avisaremos você!"
                await self.enviar_para_cliente(telefone, msg)
                self.processando.pop(telefone, None)
                return

            # --- FRETE: aguardando confirmacao do cliente ---
            if etapa == "frete_confirmar":
                opt = self._n(msg_texto.strip().lower())
                if opt in ("sim", "s", "1", "f") or opt.startswith(("sim", "s")):
                    cliente_dados = cliente_por_telefone(telefone)
                    produto = produto_por_id(conversa.get("produto_interesse_id") or 0)
                    if cliente_dados and produto:
                        ult_cot = get_ultima_cotacao(conv_id)
                        valor_frete = ult_cot.get("valor_frete", 0) if ult_cot else 0
                        xlsx_path = ult_cot.get("xlsx_path") if ult_cot else None
                        total = produto["preco"] + (valor_frete or 0)
                        link_pagamento, stripe_session_id = criar_checkout_pix_cartao(
                            nome_produto=produto["nome"],
                            valor_total=total,
                            cliente_nome=cliente_dados.get("nome", ""),
                            cliente_telefone=telefone,
                            venda_id=0,
                        )
                        venda_id = criar_venda(
                            conv_id, cliente_dados.get("id"), produto["id"],
                            produto["preco"], valor_frete=valor_frete,
                            payment_url=link_pagamento, stripe_session_id=stripe_session_id,
                            xlsx_path=xlsx_path,
                        )
                        atualizar_etapa_conversa(conv_id, "frete_aguardando_pagamento")
                        if link_pagamento:
                            await self.enviar_para_cliente(telefone,
                                f"Perfeito! Para finalizar sua compra, realize o pagamento pelo link abaixo:\n\n"
                                f"💳 {link_pagamento}\n\n"
                                f"Produto: {produto['nome']}\n"
                                f"Total: R$ {total:.2f}\n\n"
                                f"Após a confirmação do pagamento, seu pedido será processado.")
                        else:
                            await self.enviar_para_cliente(telefone,
                                f"Pedido confirmado!\nProduto: {produto['nome']}\n"
                                f"Total: R$ {total:.2f}\nObrigado pela compra!\n\n"
                                f"Entrarei em contato para finalizar o pagamento.")
                        await self.enviar_para_cliente(SEU_NUMERO,
                            f"💳 PAGAMENTO PENDENTE\n{cliente_dados.get('nome','')} - Tel: {telefone}\n"
                            f"{produto['nome']} - R$ {total:.2f}\n"
                            f"Link: {link_pagamento or 'N/D'}\nID: {venda_id}")
                        print(f"VENDA PENDENTE: {safe(cliente_dados.get('nome',''))} - {safe(produto['nome'])} - {safe(link_pagamento or 'N/D')}")
                    else:
                        await self.enviar_para_cliente(telefone, "Erro ao processar confirmação.")
                elif opt in ("nao", "não", "2", "g", "voltar") or opt.startswith(("nao", "não")):
                    ult_cot = get_ultima_cotacao(conv_id)
                    xlsx_path = ult_cot.get("xlsx_path") if ult_cot else None
                    if xlsx_path and os.path.exists(xlsx_path):
                        os.remove(xlsx_path)
                        print(f"  -> XLSX excluído: {os.path.basename(xlsx_path)}", flush=True)
                    atualizar_etapa_conversa(conv_id, "menu_principal")
                    await self.enviar_para_cliente(telefone, "Tudo bem! Se precisar de algo, estou aqui.")
                else:
                    await self.enviar_para_cliente(telefone,
                        "Não entendi. Por favor, responda:\n\n"
                        "[1] SIM - Confirmar o pedido\n"
                        "[2] NÃO - Cancelar")
                self.processando.pop(telefone, None)
                return

            # --- APRESENTACAO PROGRESSIVA DO MENU ---
            if etapa == "apresentacao_menu":
                estado = self.apresentacao_menu.get(telefone)
                if not estado:
                    self.apresentacao_menu.pop(telefone, None)
                    atualizar_etapa_conversa(conv_id, "menu_principal")
                    self.processando.pop(telefone, None)
                    return
                if msg_texto.strip().isdigit():
                    n = int(msg_texto.strip())
                    if n in estado["apresentados"]:
                        self.apresentacao_menu.pop(telefone, None)
                        produto = produto_por_id(n)
                        if produto:
                            atualizar_produto_interesse(conv_id, produto["id"])
                            self.processando.pop(telefone, None)
                            await self._iniciar_apresentacao_submenu(telefone, conv_id, produto)
                            return
                    else:
                        await self.enviar_para_cliente(telefone, f"Opção {n} ainda não foi apresentada. Aguarde as próximas opções.")
                        return
                # nao-digit: cai na conversa livre (Gemini/fallback)

            # --- APRESENTACAO PROGRESSIVA DO SUBMENU ---
            if etapa == "apresentacao_submenu":
                estado = self.apresentacao_submenu.get(telefone)
                if not estado:
                    self.apresentacao_submenu.pop(telefone, None)
                    atualizar_etapa_conversa(conv_id, "menu_principal")
                    self.processando.pop(telefone, None)
                    return
                alpha = {"a": 1, "b": 3, "c": 4, "d": 5}
                opt = msg_texto.strip().lower()
                if opt in alpha:
                    n = alpha[opt]
                    if n in estado["apresentados"]:
                        self.apresentacao_submenu.pop(telefone, None)
                        self.processando.pop(telefone, None)
                        await self._executar_opcao_submenu(telefone, conv_id, estado["produto_id"], n, nome_sidebar)
                        return
                    else:
                        await self.enviar_para_cliente(telefone, f"Opção {n} ainda não foi apresentada. Aguarde...")
                        return
                # nao-alpha: cai na conversa livre (Gemini/fallback)

            # --- SUBMENU: perguntar se deseja continuar ---
            if etapa == "submenu_continuar":
                opt = self._n(msg_texto.strip().lower())
                # Ignore long text (bot's own messages), only short user responses
                if len(opt.split()) > 3:
                    return
                ctx = self.continuar_submenu.pop(telefone, None)
                if not ctx:
                    atualizar_etapa_conversa(conv_id, "menu_principal")
                    return
                if opt in ("f", "1", "sim"):
                    produto = produto_por_id(ctx["produto_id"])
                    if produto:
                        self.processando.pop(telefone, None)
                        await self._iniciar_apresentacao_submenu(telefone, conv_id, produto)
                    return
                if opt in ("g", "2", "nao", "não", "voltar"):
                    self.apresentacao_submenu.pop(telefone, None)
                    await self.enviar_para_cliente(telefone, "Voltando ao Menu Principal...", ctx.get("nome_sidebar", ""))
                    ok = await self._iniciar_apresentacao_menu(telefone, conv_id, ctx.get("nome_sidebar", ""))
                    if ok:
                        print(f"  -> Menu reiniciado para {safe(telefone)}", flush=True)
                    return
                # Unknown input: cai na conversa livre (Gemini/fallback)
                atualizar_etapa_conversa(conv_id, "menu_principal")

            # --- SUBMENU: se estiver visualizando um produto ---
            if etapa.startswith("submenu_"):
                produto_id = int(etapa.split("_")[1])
                produto = produto_por_id(produto_id)
                if produto:
                    opt = msg_texto.strip().lower()
                    opt_map = {"a": "folder", "1": "folder", "folder": "folder",
                               "b": "foto", "3": "foto", "foto": "foto", "fotografia": "foto",
                               "c": "video", "4": "video", "video": "video", "vídeo": "video",
                               "d": "frete", "5": "frete", "frete": "frete", "cotacao": "frete", "cotaçao": "frete"}
                    acao = opt_map.get(opt)

                    if acao in ("folder", "foto", "video"):
                        if acao == "folder":
                            await self._enviar_folder(conv_id, telefone, produto)
                        elif acao == "valor":
                            resp = valor_produto(produto)
                            await self.enviar_para_cliente(telefone, resp)
                            salvar_mensagem(conv_id, "agente", resp)
                        elif acao == "foto":
                            await self._enviar_foto(conv_id, telefone, produto)
                        elif acao == "video":
                            await self._enviar_video(conv_id, telefone, produto)
                        self.processando.pop(telefone, None)
                        self.continuar_submenu[telefone] = {"conv_id": conv_id, "produto_id": produto_id, "nome_sidebar": nome_sidebar}
                        atualizar_etapa_conversa(conv_id, "submenu_continuar")
                        await self.enviar_para_cliente(telefone,
                            "Deseja mais alguma opção?\n[f] SIM - Continuar neste produto\n[g] NÃO - Voltar ao Menu Principal",
                            nome_sidebar)
                        return

                    if acao == "frete":
                        atualizar_produto_interesse(conv_id, produto["id"])
                        cli = cliente_por_telefone(telefone)
                        if cli and all([cli.get("nome"), cli.get("cpf_cnpj"), cli.get("endereco"), cli.get("cidade"), cli.get("estado"), cli.get("cep")]):
                            await self._finalizar_coleta_frete(conv_id, cli["id"], telefone, nome_sidebar)
                            return
                        atualizar_etapa_conversa(conv_id, "frete_nome")
                        await self.enviar_para_cliente(telefone,
                            f"Para solicitar o frete da {produto['nome']}, preciso de alguns dados.\n\n"
                            f"Primeiro, informe seu NOME completo:")
                        salvar_mensagem(conv_id, "agente", "Solicitando dados para frete - informe o nome:")
                        return

            # --- SELECAO DE PRODUTO: numero 1-8 -> submenu progressivo ---
            if msg_texto.strip().isdigit():
                produto = produto_por_id(int(msg_texto.strip()))
                if produto:
                    atualizar_produto_interesse(conv_id, produto["id"])
                    self.processando.pop(telefone, None)
                    await self._iniciar_apresentacao_submenu(telefone, conv_id, produto)
                    return

            # --- CONVERSA LIVRE: usa Gemini ou fallback ---
            # Pula Gemini para mensagens triviais (economiza cota)
            msg_curta = msg_texto.strip().lower().rstrip("?!.")
            triviais = {"ok", "sim", "não", "nao", "obrigado", "obrigada", "valeu",
                        "brigado", "brigada", "blz", "beleza", "tudo bem", "tudo",
                        "sim sim", "ok ok", "pode ser", "certo", "entendi", "show",
                        "legal", "perfeito", "ótimo", "otimo", "bom", "hmm", "hum",
                        "rs", "kkk", "haha", "lol", "nada", "blz blz", "tranquilo",
                        "pode deixar", "fechou", "fechado"}
            if msg_curta in triviais:
                resposta = "😊 Por nada! Estou aqui para ajudar. É só me chamar quando precisar."
                resposta_limpa = resposta
                comando = None
            else:
                # Throttle por usuário: no máximo 1 chamada Gemini a cada 15s
                ult_gem = self.ultimo_gemini.get(telefone, 0)
                if time.time() - ult_gem < 15:
                    resposta = "Estou processando sua solicitação, aguarde um momento..."
                    resposta_limpa = resposta
                    comando = None
                else:
                    historico = get_historico_conversa(conv_id, limite=10)
                    try:
                        resposta = await asyncio.to_thread(gerar_resposta, historico)
                        self.ultimo_gemini[telefone] = time.time()
                    except Exception as e:
                        print(f"[GROQ] {safe(e)}")
                        agora_fb = time.time()
                        ult_fb = self.ultimo_fallback.get(telefone, 0)
                        if agora_fb - ult_fb > 3600:
                            resposta = resposta_fallback(historico)
                            self.ultimo_fallback[telefone] = agora_fb
                        else:
                            resposta = "Desculpe, estou temporariamente offline. Tente novamente mais tarde."

                    comando = extrair_comando(resposta)
                    resposta_limpa = limpar_resposta(resposta)

            if resposta_limpa:
                await self.enviar_para_cliente(telefone, resposta_limpa, nome_sidebar)
                salvar_mensagem(conv_id, "agente", resposta_limpa)
            if comando:
                await self.executar_comando(comando, conv_id, cliente_id, telefone, remetente, nome_sidebar)

        except Exception as e:
            print(f"[ERRO processar] {safe(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.processando.pop(telefone, None)

    async def executar_comando(self, comando: dict, conv_id, cliente_id, telefone, remetente, nome_sidebar: str = ""):
        acao = comando["acao"]
        if acao == "enviar_midia":
            produto = next((p for p in PRODUTOS if p["id"] == comando["produto_id"]), None)
            if produto:
                if comando["tipo"] == "foto":
                    await self._enviar_foto(conv_id, telefone, produto)
                elif comando["tipo"] == "video":
                    await self._enviar_video(conv_id, telefone, produto)
                elif comando["tipo"] == "folder":
                    await self._enviar_folder(conv_id, telefone, produto)
                atualizar_produto_interesse(conv_id, produto["id"])

        elif acao == "solicitar_frete":
            await self._executar_frete(conv_id, cliente_id, telefone, remetente)

        elif acao == "venda_confirmada":
            produto = next((p for p in PRODUTOS if p["id"] == comando["produto_id"]), None)
            if not produto:
                return
            venda_id = criar_venda(conv_id, cliente_id, produto["id"], produto["preco"])
            atualizar_etapa_conversa(conv_id, "fechada")
            total = comando["valor_total"]
            # Gera link de pagamento Stripe (Pix + Cartão)
            link_pagamento, stripe_session_id = criar_checkout_pix_cartao(
                nome_produto=produto["nome"],
                valor_total=total,
                cliente_nome=comando["cliente_nome"],
                cliente_telefone=telefone,
                venda_id=venda_id,
            )
            if link_pagamento:
                from database import get_connection
                conn = get_connection()
                conn.execute("UPDATE vendas SET payment_url=?, stripe_session_id=? WHERE id=?",
                             (link_pagamento, stripe_session_id, venda_id))
                conn.commit()
                conn.close()
                await self.enviar_para_cliente(telefone,
                    f"Venda confirmada!\nProduto: {produto['nome']}\n"
                    f"Total: R$ {total:.2f}\n\n"
                    f"💳 Link para pagamento (Pix ou Cartão):\n{link_pagamento}")
            else:
                await self.enviar_para_cliente(telefone,
                    f"Venda confirmada!\nProduto: {produto['nome']}\n"
                    f"Total: R$ {total:.2f}\nObrigado!")
            await self.enviar_para_cliente(SEU_NUMERO,
                f"VENDA!\n{comando['cliente_nome']} - Tel: {telefone}\n"
                f"{produto['nome']} - R$ {total:.2f}\n"
                f"Link: {link_pagamento or 'N/D'}\nID: {venda_id}")
            print(f"VENDA REGISTRADA: {safe(comando['cliente_nome'])} - {safe(produto['nome'])}")

    def extrair_valor_frete(self, texto: str) -> float:
        padroes = [
            r"(?:VALOR\s*(?:DO\s*)?FRETE|FRETE)\s*:?\s*(?:R\$)?\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)",
            r"(?:R\$)\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)",
            r"(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))\s*(?:reais|R\$)?",
        ]
        for p in padroes:
            m = re.search(p, texto, re.IGNORECASE)
            if m:
                val = m.group(1)
                val = re.sub(r"[^\d,.]", "", val)
                if "," in val and "." in val:
                    if val.rindex(",") > val.rindex("."):
                        val = val.replace(".", "").replace(",", ".")
                    else:
                        val = val.replace(",", "")
                elif "," in val:
                    val = val.replace(".", "").replace(",", ".")
                return float(val)
        return 0.0

    def extrair_prazo(self, texto: str) -> str | None:
        padroes = [
            r"(\d+[_\s-]*dias?\s*úteis?)",
            r"(\d+[_\s-]*dias?\s*uteis)",
            r"(\d+[_\s-]*dias?\s*corridos?)",
            r"(\d+[_\s-]*dias?)",
        ]
        for p in padroes:
            m = re.search(p, texto, re.IGNORECASE)
            if m:
                return m.group(0)
        return None

    def extrair_protocolo_transportadora(self, texto: str) -> str | None:
        padroes = [
            r"Protocolo\s+Transportadora:\s*(\S+)",
            r"Protocolo[:\s]+\s*(\S+)",
            r"Prot[.:]?\s*(?:transp[.:]?)?\s*(\S+)",
            r"(?:COD[.:]?|PROT(?!OCOL)[.:]?|PEDIDO[.:]?)\s*(\d{4,})",
        ]
        for p in padroes:
            m = re.search(p, texto, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if len(val) < 3 or val.startswith(("VALOR", "PRAZO", "Protocolo")):
                    continue
                return val
        return None

    async def parar(self):
        if self.frete_monitor_task:
            self.frete_monitor_task.cancel()
            try:
                await self.frete_monitor_task
            except asyncio.CancelledError:
                pass
            self.frete_monitor_task = None
        if self.fila_worker_task:
            self.fila_worker_task.cancel()
            try:
                await self.fila_worker_task
            except asyncio.CancelledError:
                pass
            self.fila_worker_task = None
        if self.transport_page:
            try:
                await self.transport_page.close()
            except:
                pass
            self.transport_page = None
        if self.context:
            await self.context.close()
        if self.playwright:
            await self.playwright.stop()


async def main():
    bot = WhatsAppBot()
    try:
        ok = await bot.iniciar()
        if ok:
            await bot.escutar_mensagens()
    except KeyboardInterrupt:
        print("\nEncerrando...")
    finally:
        await bot.parar()


if __name__ == "__main__":
    asyncio.run(main())
